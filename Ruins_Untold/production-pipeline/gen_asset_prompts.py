#!/usr/bin/env python3
"""
Ruins Untold — Episode Asset Prompt Generator v4
Reads:  {PROJ}/scripts/voiceover_transcript.srt  (Whisper output)
Writes: {PROJ}/scripts/timed_transcript.txt
        {PROJ}/scripts/asset_prompts.json
        {PROJ}/scripts/image_prompts.xlsx
        {PROJ}/scripts/video_prompts.xlsx
        {PROJ}/scripts/music_prompts.xlsx
        {PROJ}/scripts/sfx_prompts.xlsx
        {PROJ}/scripts/media_placement.json
        {PROJ}/scripts/assemble.sh

v4 CHANGE: All 224 visual prompts are written directly from the timed
transcript. Each cue has a manually-crafted description tied to exactly
what the narrator is saying at that moment. No keyword matching, no
entity extraction, no scene_type field, no static variant pools.

Usage:
    python3 gen_asset_prompts.py --proj /path/to/episode/dir [--topic "Title"]

Channel config:
    /Users/jneal/Antigravity/n8nYT_Script/ruins_untold_channel_config.json
"""

import argparse, json, re, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CHANNEL_CONFIG = "/Users/jneal/Antigravity/n8nYT_Script/ruins_untold_channel_config.json"

# ─────────────────────────────────────────────────────────────────────────────
# Args
# ─────────────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--proj",  required=True, help="Episode project directory")
parser.add_argument("--topic", default="",    help="Episode topic/title")
args = parser.parse_args()

PROJ    = args.proj.rstrip("/")
SCRIPTS = f"{PROJ}/scripts"
TOPIC   = args.topic or os.path.basename(PROJ)

SRT_FILE   = f"{SCRIPTS}/voiceover_transcript.srt"
TRANSCRIPT = f"{SCRIPTS}/timed_transcript.txt"

# ─────────────────────────────────────────────────────────────────────────────
# Load channel config
# ─────────────────────────────────────────────────────────────────────────────
with open(CHANNEL_CONFIG) as f:
    cfg = json.load(f)

STYLE    = cfg["style_constants"]["style"]
PALETTE  = cfg["style_constants"]["palette"]
LIGHTING = cfg["style_constants"]["lighting"]
NEG      = cfg["style_constants"]["negative"]
NEG_VID  = cfg["style_constants"]["negative_vid"]

IMAGE_INTERVAL = cfg["timing"]["image_interval_seconds"]   # 7
VIDEO_EVERY    = cfg["timing"]["video_every_n_cues"]        # 4
FADE_DUR       = cfg["timing"]["crossfade_duration_seconds"]# 0.5
J_CUT_OFFSET   = cfg["timing"].get("j_cut_offset_seconds", 0.0)  # audio leads video by this many seconds

THUMB_CFG      = cfg.get("thumbnail", {})

PINNED_CLIPS  = cfg.get("pinned_clips",  [])
PINNED_IMAGES = cfg.get("pinned_images", [])

def check_pinned(text):
    tl = text.lower()
    for entry in PINNED_CLIPS + PINNED_IMAGES:
        for phrase in entry["trigger_phrases"]:
            if entry["match"] == "contains" and phrase in tl:
                return entry
            if entry["match"] == "exact" and phrase == tl.strip():
                return entry
    return None

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def srt_ts_to_sec(ts):
    ts = ts.replace(",", ".")
    h, m, rest = ts.split(":")
    return int(h)*3600 + int(m)*60 + float(rest)

def sec_to_ts(sec):
    sec = max(0.0, sec)
    return f"{int(sec)//60:02d}:{int(sec)%60:02d}"

def ts_key(ts_str):
    p = ts_str.split(":")
    return int(p[0])*60 + int(p[1])

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Parse SRT → timed_transcript.txt
# ─────────────────────────────────────────────────────────────────────────────
with open(SRT_FILE) as f:
    raw = f.read()

blocks = []
for block in re.split(r'\n\n+', raw.strip()):
    lines_raw = [l.strip() for l in block.strip().split('\n') if l.strip()]
    if len(lines_raw) < 3:
        continue
    m = re.match(r'(\S+)\s+-->\s+(\S+)', lines_raw[1])
    if not m:
        continue
    start_sec = srt_ts_to_sec(m.group(1))
    text = ' '.join(lines_raw[2:]).strip().lstrip()
    if text:
        blocks.append({"sec": start_sec, "text": text})

# Merge short fragments
merged = []
buf_sec, buf_text = None, []
for b in blocks:
    word_count = len(b["text"].split())
    if buf_sec is None:
        buf_sec, buf_text = b["sec"], [b["text"]]
    elif word_count <= 3 and buf_text:
        buf_text.append(b["text"])
    else:
        merged.append({"sec": buf_sec, "text": " ".join(buf_text)})
        buf_sec, buf_text = b["sec"], [b["text"]]
if buf_text:
    merged.append({"sec": buf_sec, "text": " ".join(buf_text)})

ACTUAL_DUR = blocks[-1]["sec"]

with open(TRANSCRIPT, "w") as f:
    for m in merged:
        f.write(f"[{sec_to_ts(m['sec'])}] {m['text']}\n")

print(f"✓ timed_transcript.txt  {len(merged)} lines  {sec_to_ts(merged[0]['sec'])} → {sec_to_ts(merged[-1]['sec'])}")
lines = merged

def nearest_text(sec):
    return min(lines, key=lambda l: abs(l["sec"] - sec))["text"]

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — PROMPT_LIST
# 224 entries indexed by cue_n (0-223).
# Cue schedule: IMAGE_INTERVAL=7s, VIDEO_EVERY=4.
# cue_n=0 → 0s, cue_n=1 → 7s, ..., cue_n=223 → 1561s
# Videos fire at cue_n: 4,8,12,16,20,24,28,32,36,40,44,48,52,56,60,64,68,72,
#                       76,80,84,88,92,96,100,104,108,112,116,120,124,128,132,
#                       136,140,144,148,152,156,160,164,168,172,176,180,184,
#                       188,192,196,200,204,208,212,216,220
# cue_n=13 (91s) is intercepted as PINNED VIDEO — prompt slot unused.
# Each entry: (description, mood)
# ─────────────────────────────────────────────────────────────────────────────
PROMPT_LIST = [
    # cue 0 — 0s — "There is a document, a real document..."
    ("Close-up of the cover of the 1868 Smithsonian Institution Annual Report — thick cloth-bound institutional volume, gold-stamped Smithsonian seal embossed on the deep ochre cover, resting on a dark mahogany archive desk",
     "archival, revelatory, institutional, weighty"),
    # cue 1 — 7s — "not by a tabloid"
    ("Sensationalist 19th-century yellow journalism tabloid with bold gothic headline beside a sober cloth-bound Smithsonian Annual Report — the two documents laid side by side on a reading table, gaslit amber light contrasting the sources",
     "comparative, period-authentic, investigative"),
    # cue 2 — 14s — "It describes the excavation of a human skeleton...burial mound in Ohio"
    ("Wide period excavation scene — Ohio burial mound, mid-1800s, workers in period field attire digging a deep trench into an earthen mound, dark ochre soil walls exposed, skeletal remains partially uncovered at the base of the pit, raking afternoon light",
     "archaeological, ancient, revelatory"),
    # cue 3 — 21s — "measured by a trained field agent...7 feet, 2 inches"
    ("Bureau of Ethnology field agent in Victorian attire kneeling beside fully excavated skeletal remains in an Ohio burial pit — extending a measuring rod alongside the bones, the exceptional length of the skeleton unmistakable against the rod",
     "forensic, shocking, archaeological, period-authentic"),
    # cue 4 — 28s — VIDEO — "That document still exists...dated 1868"
    ("Slow push into an open 1868 Smithsonian Annual Report on an archive desk — the camera drifts close over yellowed foxed pages, field agent measurement notation and the year 1868 coming into focus in period institutional typography",
     "archival, revelatory, documentary"),
    # cue 5 — 35s — "Smithsonian's own annual report series, archived and partially digitized"
    ("Long row of Smithsonian Bureau of Ethnology Annual Report volumes on dark archive shelves — spines labeled by year from the 1860s through the 1890s, warm amber reading room light, dust motes visible in the still air",
     "archival, institutional, investigative"),
    # cue 6 — 42s — "So where are the skeletons? Where are any of them?"
    ("Empty excavation trench in Ohio mound country — a deep pit with dark ochre soil walls, freshly disturbed earth at the base, but the trench completely empty: no remains, no artifacts, no evidence of what was once there",
     "ominous, conspicuous absence, archaeological"),
    # cue 7 — 49s — "The accepted history of pre-Columbian North America tells us that"
    ("Interior of a 19th-century natural history museum — glass cases containing human skeletal remains of normal stature, brass placards in institutional typography, warm gaslit interior, the curated authority of the official narrative",
     "authoritative, institutional, period-authentic"),
    # cue 8 — 56s — VIDEO — "no human remains exceeding normal height ranges have ever been found"
    ("Slow tracking shot through a 19th-century natural history museum hall — display cases on either side containing skeletal remains of ordinary stature, camera moves steadily forward through the ordered institutional certainty of the official position",
     "authoritative, institutional, investigative"),
    # cue 9 — 63s — "giant skeleton reports are hoaxes, newspaper fabrications, yellow journalism"
    ("Stack of sensationalist 19th-century American newspapers piled on a reading table — lurid bold headlines about giant skeleton discoveries, exaggerated gothic typefaces, yellow and foxed paper, the visible machinery of yellow journalism",
     "dismissive, period-authentic, documentary"),
    # cue 10 — 70s — "stories told by people trying to sell papers"
    ("19th-century newspaper printing press in full operation — period wooden press, ink rollers, aproned workers feeding broadsheet paper, warm industrial gaslight, the commercial engine driving sensationalist coverage",
     "period-authentic, industrial, contextual"),
    # cue 11 — 77s — "That is a very clean explanation. It would be more convincing if..."
    ("Immaculately ordered academic desk — 19th-century institutional setting, open textbook with illustrated charts, professor's notes in neat period penmanship, an arrangement too clean, too settled, something conspicuously missing from the tidy picture",
     "deceptive neatness, institutional, investigative"),
    # cue 12 — 84s — VIDEO — "the Smithsonian had not written the opposite down in its own records under its own name"
    ("Slow tilt-up along a shelf of Smithsonian Annual Report volumes — camera begins at the base spine reading 1862 and tilts slowly upward past 1868, 1873, 1882, 1889, two decades of the institution's own name stamped on every spine",
     "accumulating weight, archival, revelatory"),
    # cue 13 — 91s — PINNED VIDEO — "Welcome to Ruins Untold"
    # This slot is intercepted by the pinned channel intro clip. Description unused.
    ("[PINNED CHANNEL ASSET]", "channel intro"),
    # cue 14 — 98s — "I want to be precise about what we are and are not claiming here"
    ("Methodical researcher's document table — primary source Bureau of Ethnology reports organized in labeled folders, red pen annotations marking specific passages, magnifying glass on an open page, single overhead lamp, deliberate scholarly precision",
     "investigative, measured, methodical"),
    # cue 15 — 105s — "buried under so many layers of bad faith reporting"
    ("Archaeological cross-section of soil strata beneath a burial mound — the excavation trench wall showing horizontal bands of earth representing centuries of accumulation, the deepest layer still sealed below, the actual evidence buried under layers",
     "metaphorical, geological, investigative"),
    # cue 16 — 112s — VIDEO — "sensationalist thumbnails and outright fabrication...actual evidence almost impossible to see"
    ("Slow pan across a cluttered wall of 19th-century sensationalist newspaper clippings pinned edge-to-edge — lurid giant skeleton headlines, exaggerated engravings — camera moves left to right before pausing on a single sober Bureau field report isolated in the corner",
     "chaotic, contrast, investigative, revelatory"),
    # cue 17 — 119s — "So let us clear the ground"
    ("Single authentic Bureau of Ethnology field report emerging isolated from a scattered pile of newspaper clippings — the official institutional document spotlit clearly against the clutter, institutional letterhead visible at top",
     "clarifying, documentary, focused"),
    # cue 18 — 126s — "The mainstream position is this"
    ("Formal 19th-century academic lecture hall — mahogany podium with institutional crest, empty rows of period wooden seating beyond, single gaslit lamp illuminating the speaker's position, the setting where official positions are proclaimed",
     "authoritative, institutional, period-authentic"),
    # cue 19 — 133s — "newspapers throughout the 19th century are almost universally journalistic exaggeration"
    ("Collection of 19th-century American newspaper mastheads spread flat — multiple regional publications from Ohio, Virginia, Illinois, different typographic styles, foxed paper edges, the print media landscape of the mound builder era",
     "period-authentic, journalistic, documentary"),
    # cue 20 — 140s — VIDEO — "The Cardiff Giant hoax of 1869, a carved gypsum figure presented as a petrified 10-foot man"
    ("Slow push toward the Cardiff Giant gypsum figure lying in a shallow pit — camera approaches from the feet end, the enormous carved stone figure filling the frame, period onlookers visible at the pit edges, rural New York state, 1869",
     "cautionary, period-authentic, deceptive"),
    # cue 21 — 147s — "is the centerpiece of that argument. Hoaxes did happen."
    ("Extreme close-up of the Cardiff Giant's carved stone face — crude facial features of the hoax figure, rough-hewn gypsum texture, the forgery quality undeniable at close range, a cautionary example",
     "revealing, cautionary, period-authentic"),
    # cue 22 — 154s — "The 19th century press was not exactly rigorous"
    ("Chaotic 1870s newsroom interior — crowded editorial office with typesetters at cases, papers scattered across every surface, editor at a cluttered desk under a gas lamp, the disorder of speed over accuracy",
     "period-authentic, chaotic, contextual"),
    # cue 23 — 161s — "the Hopewell and Adena mound building cultures of Ohio, Indiana, and Illinois"
    ("Aerial view of Hopewell earthwork complex — the geometric precision of octagonal and circular enclosures spreading across Ohio grassland, dawn mist low in the valleys between earthen walls, no modern elements",
     "ancient, monumental, geometric, civilizational"),
    # cue 24 — 168s — VIDEO — "burial mounds systematically excavated throughout the late 19th century"
    ("Tracking shot across an active 19th-century mound excavation — period workers with shovels and brushes working systematically along a trench, wooden survey stakes at measured intervals, camera moves slowly left revealing the scale of the systematic program",
     "archaeological, methodical, period-authentic"),
    # cue 25 — 175s — "ancestral to modern Native American populations. Biological evidence from those mounds"
    ("Authentic Hopewell culture grave goods in museum context — copper ornaments, shell bead necklaces, stone tools arranged respectfully in an institutional display case, the genuine material culture of the burial mound peoples",
     "cultural, respectful, archaeological, period-authentic"),
    # cue 26 — 182s — "fully consistent with known human variation"
    ("19th-century physical anthropology illustration — scientific diagram of skeletal measurements and human variation, comparative column format, period scientific engraving style in an institutional publication",
     "scientific, period-authentic, comparative"),
    # cue 27 — 189s — "claims of giant remains reflect misidentification, exaggeration, or fraud. The official position."
    ("Heavily stamped official Smithsonian institutional document — thick institutional letterhead, authoritative period typography, the language of a settled conclusion, the formal weight of the official position",
     "authoritative, final, institutional, closed"),
    # cue 28 — 196s — VIDEO — "The problem is what the Smithsonian's own agents wrote down in the field before that position became policy"
    ("Push into a stack of pre-1894 Bureau of Ethnology field reports — camera drifts slowly toward handwritten field notations on yellowed paper, dates in the 1870s and early 1880s visible at the tops of pages, the pre-policy record",
     "archival, investigative, revelatory"),
    # cue 29 — 203s — "Bureau of Ethnology, the Smithsonian's own archaeological arm, created in 1879"
    ("1879 Congressional mandate establishing the Bureau of Ethnology — formal government document on aged institutional paper, congressional seal visible, the founding charter of the Smithsonian's archaeological arm",
     "institutional, founding, period-authentic"),
    # cue 30 — 210s — "produced field reports, annual reports, documented findings under names of real scientists"
    ("Array of handwritten Bureau of Ethnology field reports spread flat — multiple pages with different agents' handwriting, real names at the top of each notation, dates ranging 1879 to 1889, the institutional record of real scientists",
     "archival, documentary, investigative"),
    # cue 31 — 217s — "working real sites. Those reports describe things the current official position says do not exist."
    ("Split composition — on the left, an 1882 Bureau field report with specific measurement notations; on the right, a modern institutional position statement — the two documents laid flat side by side, the contradiction visible",
     "contradictory, investigative, revelatory"),
    # cue 32 — 224s — VIDEO — "That gap between what their own agents documented and what the institution acknowledges"
    ("Slow pan across a documentary timeline — handwritten Bureau field report dates 1868, 1873, 1882 on the left side of a physical gap in the documents, the 1894 official conclusion on the right — camera moves across the void between the two",
     "gap, investigative, revelatory, ominous"),
    # cue 33 — 231s — "To understand what was found, you need to understand what was being excavated"
    ("Panoramic view of the American interior in the mid-1800s — vast Ohio River valley autumn landscape, burial mounds visible as low grassy rises across the plain, the territory of the mound builders, no modern elements",
     "ancient, geographic, contextual, vast"),
    # cue 34 — 238s — "mid to late 19th century, the American interior was experiencing something like a gold rush"
    ("Period scene of American westward movement through the Ohio Valley — settlers with wagons moving across the river valley landscape, burial mounds visible on the ridgelines behind them, the energy of 19th-century expansion",
     "historical, period-authentic, expansive"),
    # cue 35 — 245s — "except instead of gold, people were finding mounds, thousands of them"
    ("Aerial perspective over the Ohio interior showing dozens of burial mounds — circular and linear grassy rises across the landscape as far as the eye can see, the overwhelming number of the mound builder earthworks",
     "ancient, vast, overwhelming, geographic"),
    # cue 36 — 252s — VIDEO — "Earth and burial mounds, effigy mounds, platform mounds, scattered across Ohio, Indiana, Illinois, Kentucky"
    ("Aerial pull-back revealing burial mounds stretching across the autumn Ohio landscape — camera rises slowly to show the geographic breadth of the mound builder territory, mounds extending to the horizon in every direction",
     "vast, ancient, civilizational, geographic"),
    # cue 37 — 259s — "Earth and burial mounds, effigy mounds, platform mounds"
    ("Three mound types in one composition — a conical burial mound, a flat platform mound, and the curved outline of an effigy mound form visible from low aerial angle, the distinct architectural vocabulary of the mound builders",
     "archaeological, comparative, ancient, educational"),
    # cue 38 — 266s — "Hopewell culture, peaked roughly between 100 BCE and 500 CE"
    ("Historical timeline in period cartographic style — a hand-annotated timeline bar spanning 100 BCE to 500 CE, the Hopewell culture peak highlighted, overlaid on a map of the Ohio Valley, period compass rose in the corner",
     "historical, cartographic, period-authentic"),
    # cue 39 — 273s — "built mound complexes of extraordinary geometric precision. Newark, Ohio. Chillicothe. Mound City."
    ("Aerial view of the Newark Earthworks — the vast octagonal and circular geometric enclosures stretching across the Ohio plain, the extraordinary mathematical precision of the ancient engineering visible from elevation, morning mist in the low places",
     "monumental, geometric, ancient, civilizational"),
    # cue 40 — 280s — VIDEO — "Some of these earthworks enclosed areas of hundreds of acres, with geometric alignments to astronomical events"
    ("Aerial pull-back from the center of the Newark Great Circle earthwork — camera rises slowly to reveal the circular earthen walls enclosing hundreds of acres, the perfect geometric form of the ancient structure against the Ohio landscape",
     "monumental, ancient, geometric, awe-inspiring"),
    # cue 41 — 287s — "that we are still mapping today. Who built them?"
    ("LiDAR survey elevation map of a Hopewell earthwork complex — the mathematical precision of the mound geometry rendered in subtle color gradients, earthwork details invisible to the naked eye revealed by the technology",
     "technological, revealing, precise, investigative"),
    # cue 42 — 294s — "That question drove American anthropology and archaeology for the better part of a century"
    ("19th-century American archaeological society meeting — period illustration of learned men in formal attire gathered around a table covered in mound maps, intense scholarly debate implied by their postures and gestures",
     "historical, scholarly, period-authentic, investigative"),
    # cue 43 — 301s — "The early answer was that the mounds were built by a lost, advanced civilization"
    ("Romanticized 19th-century engraving of the 'Mound Builders' civilization — published in a period popular science journal, depicting monumental earthwork construction, idealized figures in ceremonial attire, the mythology of the lost race",
     "mythological, period-authentic, speculative, romantic"),
    # cue 44 — 308s — VIDEO — "sometimes called the mound builders, that had been destroyed or displaced before European contact"
    ("Slow dolly through a period reconstruction scene — grand mound builder settlement at its peak, earthworks under construction, costumed figures in ceremonial procession, camera drifts through the imagined civilization that drove a century of debate",
     "reconstructed, ancient, mysterious, civilizational"),
    # cue 45 — 315s — "Thomas Jefferson, who excavated a mound on his Virginia property in 1784"
    ("Thomas Jefferson in period attire supervising the systematic excavation of a Virginia burial mound — workers with period tools at the trench edge, Jefferson making careful notes in a field journal, the first scientific approach to the mound question, 1784",
     "historical, pioneering, period-authentic, methodical"),
    # cue 46 — 322s — "His conclusion was careful and somewhat ahead of its time"
    ("Jefferson's 1784 mound excavation field journal open to his notes — careful period penmanship, systematic observations, measured and precise scientific language, the founding document of American mound archaeology",
     "scholarly, historical, pioneering, archival"),
    # cue 47 — 329s — "The mounds were built by indigenous peoples, possibly ancestral to modern tribes"
    ("Authentic Hopewell culture copper ceremonial ornament — green-patinated hammered copper headdress element, photographed in warm directional light against a dark background, clearly indigenous craftsmanship of high skill",
     "cultural, authentic, indigenous, ancient"),
    # cue 48 — 336s — VIDEO — "by the mid-1800s, the popular imagination had run considerably further than Jefferson"
    ("Pan across a collection of 1840s-1860s popular publications speculating about Mound Builders — romanticized engravings, 'Lost Race' and 'Ancient Civilization' headlines visible on covers, camera moves left across the spread of mythologizing publications",
     "mythological, period-authentic, popular imagination"),
    # cue 49 — 343s — "The mound builders had become, in the public mind, a lost race — Israelite, Phoenician, Welsh"
    ("19th-century illustrated newspaper page showing speculative Mound Builder origins — engraving depicting Phoenician traders landing in America beside indigenous peoples, bold caption text beneath, the mythology in full popular expression",
     "mythological, speculative, period-authentic"),
    # cue 50 — 350s — "America's origin myth needed a dramatic backstory"
    ("Period American popular history book cover — dramatic illustrated cover proclaiming 'The Lost Race of America,' bold Victorian graphic design, the commercialization of American prehistory mythology",
     "mythological, commercial, period-authentic"),
    # cue 51 — 357s — "This is the context in which the Bureau of Ethnology was created in 1879"
    ("1879 Congressional Record page documenting the legislation creating the Bureau of Ethnology — formal government typography on period paper, the institutional founding moment in its official document form",
     "institutional, founding, period-authentic, governmental"),
    # cue 52 — 364s — VIDEO — "Its mandate, under founding director John Wesley Powell, was to settle the mound builder question"
    ("Push into a formal Victorian portrait of John Wesley Powell at his Bureau of Ethnology desk — institutional setting, Bureau founding charter documents spread before him, camera slowly approaches the authoritative figure charged with settling the mound question",
     "authoritative, institutional, biographical, period-authentic"),
    # cue 53 — 371s — "to conduct systematic scientific excavation, to produce a definitive report"
    ("Systematic archaeological excavation grid over a burial mound — wooden survey stakes at precisely measured intervals forming a scientific grid, period measuring ribbons, the methodology of systematic science being applied to the mounds",
     "methodical, scientific, period-authentic, archaeological"),
    # cue 54 — 378s — "That report arrived in 1894"
    ("The 1894 Mound Explorations report — heavy cloth-bound US government publication on an archive shelf, 'Mound Explorations' and 'Bureau of Ethnology' on the spine, the institutional weight of the definitive publication",
     "institutional, final, authoritative, archival"),
    # cue 55 — 385s — "It was written primarily by Cyrus Thomas"
    ("Portrait-style framing of Cyrus Thomas — Victorian-era scientist in period frock coat, seated at a field desk surrounded by open notebooks and specimen jars, pen in hand over a measurement notation, the man who wrote the definitive report",
     "biographical, period-authentic, scholarly, complex"),
    # cue 56 — 392s — VIDEO — "its conclusion was unambiguous. The mounds were built by the ancestors of modern Native Americans."
    ("Slow pan across the 1894 Mound Explorations report conclusion pages — camera drifts across the clear institutional text declaring the mound builder question settled, 'ancestors of modern Native Americans' visible in the period typography",
     "definitive, institutional, closing, authoritative"),
    # cue 57 — 399s — "There was no lost race. The mystery, officially, was solved."
    ("The 1894 Mound Explorations report shut closed on an archive shelf — the spine facing camera, dusty adjacent volumes, the finality of the officially solved question implied by the closed book",
     "final, closed, institutional, settled"),
    # cue 58 — 406s — "Case closed. Settled science. Except, between 1868 and 1889..."
    ("Chronological ledger or calendar with the years 1868 through 1889 highlighted — the specific 21-year window before the definitive report, annotated in red as the crucial period, an institutional timeline with the anomalous window marked",
     "investigative, temporal, revelatory"),
    # cue 59 — 413s — "the Bureau's own field agents were submitting reports, findings from the ground"
    ("Stack of pre-1894 Bureau field reports from different years — 1868, 1873, 1882 clearly visible on cover pages, different handwriting styles on each, the accumulating documentary record of the pre-settlement period",
     "archival, documentary, accumulating"),
    # cue 60 — 420s — VIDEO — "some of what they were finding did not fit cleanly into the narrative that would eventually be settled"
    ("Push into a Bureau of Ethnology field report — camera drifts slowly toward a specific passage where measurement notations appear that do not conform to the later official conclusion, the anomalous text coming into focus",
     "investigative, revelatory, archival, ominous"),
    # cue 61 — 427s — "those annual reports are real documents. They exist."
    ("Physical Smithsonian Annual Report volume held open in both hands — the tangible physical reality of the document, institutional letterhead clearly visible, the material existence of the primary source emphasized",
     "tangible, documentary, archival, investigative"),
    # cue 62 — 434s — "some of what they describe has never been adequately explained"
    ("Specific passage in a Bureau field report ringed in red pen — the anomalous measurement notation circled clearly, surrounding routine text providing context, the one entry that defies the official explanation",
     "investigative, focused, revelatory"),
    # cue 63 — 441s — "Not a newspaper headline. Not a second-hand account."
    ("Authentic Bureau of Ethnology primary source field note close-up — handwritten notation on period government paper, the unmistakable format of an official institutional record, no newspaper format present",
     "primary source, institutional, documentary, definitive"),
    # cue 64 — 448s — VIDEO — "The Smithsonian Annual Report for 1868 includes field notes from an excavation of burial mounds in Ohio"
    ("Slow push into the open 1868 Smithsonian Annual Report — camera begins wide on the open institutional volume and slowly pushes toward the Ohio excavation section, the field notes coming into focus against the foxed yellowed pages",
     "archival, investigative, revelatory, documentary"),
    # cue 65 — 455s — "documented by a bureau field agent, submitted under the institutional letterhead of the Smithsonian itself"
    ("Close-up of the 1868 field report header — Bureau of Ethnology institutional letterhead clearly visible at the top of the document, the official Smithsonian stamp of authority, a real agent's name below the heading",
     "institutional, documentary, archival, authoritative"),
    # cue 66 — 462s — "The notation describes skeletal remains recovered from a mound context"
    ("Field agent's handwritten notation on period paper — the careful scientific language of a trained observer recording skeletal remains, Ohio mound site heading above, institutional language throughout",
     "archival, forensic, documentary, period-authentic"),
    # cue 67 — 469s — "The recorded measurement. Seven feet, two inches."
    ("Extreme close-up of the specific measurement entry — the handwritten '7 feet, 2 inches' in period ink, the central anomalous fact of the document isolated in the frame, the number that shouldn't exist",
     "shocking, forensic, revelatory, focused"),
    # cue 68 — 476s — VIDEO — "you may be thinking measurement error, bone displacement, burial context"
    ("Low tracking shot through excavation trench interior — camera moves slowly along the base of a deep pit, dark soil walls on either side, scattered bone fragments visible in situ, the legitimate methodological uncertainty made visible",
     "forensic, archaeological, ambiguous, investigative"),
    # cue 69 — 483s — "Incomplete articulation leading to overcalculation. These are legitimate methodological concerns."
    ("Period scientific diagram of bone displacement and articulation in a burial context — the legitimate technical concerns illustrated, incomplete skeletal articulation shown in a field notebook illustration, measured scientific inquiry",
     "scientific, legitimate, forensic, contextual"),
    # cue 70 — 490s — "But the 1868 Ohio report is not alone"
    ("Three Bureau field reports spread side by side on a document table — 1868 Ohio, 1873 West Virginia, 1882 Illinois — each with measurement notations visible, the pattern of three separate finds emerging across three documents",
     "pattern-revealing, investigative, accumulating, revelatory"),
    # cue 71 — 497s — "the pattern of what was found, and then what happened to it"
    ("Investigation board — documents pinned to a surface with red thread connecting 1868, 1873, and 1882 Bureau reports to a central question mark, each thread tracing the path from discovery to disappearance",
     "investigative, pattern-revealing, conspiratorial"),
    # cue 72 — 504s — VIDEO — "Let us go through what the records actually show, carefully, source by source"
    ("Slow pan left-to-right across three Bureau field reports laid flat in chronological sequence — 1868, 1873, 1882 — camera moves methodically across each document, the source-by-source examination made visual",
     "methodical, documentary, investigative, archival"),
    # cue 73 — 511s — "The 1868 Smithsonian Annual Report, Ohio Mound Excavation. Field Agent notation."
    ("Open 1868 Smithsonian Annual Report to the Ohio Mound Excavation section — Ohio heading clearly readable, a real field agent's name visible below, warm amber reading lamp light on the foxed institutional pages",
     "archival, primary source, documentary, warm"),
    # cue 74 — 518s — "Skeletal remains measuring seven feet, two inches. Recorded in the same institutional document series"
    ("The 1868 measurement entry in its institutional context — '7 feet, 2 inches' visible among a column of routine Bureau findings, uncontroversial scientific entries above and below, the anomaly embedded in the normal record",
     "contextual, revelatory, institutional, archival"),
    # cue 75 — 525s — "that archives the bureau's routine findings. Not a newspaper. Not a local legend."
    ("Full spread of a 1868 Annual Report page — agricultural surveys, weather observations, routine Bureau findings filling the columns, the skeletal measurement entry appearing among unremarkable institutional records",
     "institutional, documentary, contextual, archival"),
    # cue 76 — 532s — VIDEO — "A bureau field report. We will come back to what this means."
    ("Pull back from the 1868 Annual Report — camera begins tight on the measurement notation and slowly pulls back to reveal the full institutional page, then the full open volume, placing the anomalous entry in its official documentary context",
     "contextual, archival, revealing, institutional"),
    # cue 77 — 539s — "The 1873 Smithsonian Annual Report documents excavations from a burial mound site in West Virginia"
    ("Open 1873 Smithsonian Annual Report to the West Virginia Excavation section — the new year and new state heading clearly readable, a different agent's name below, institutional consistency across the two reports",
     "archival, documentary, primary source, investigative"),
    # cue 78 — 546s — "Multiple skeletal remains. More than one individual."
    ("Wide shot into an active 1873 West Virginia excavation trench — multiple distinct burial contexts visible in the dark soil, more than one set of remains at different positions in the pit, period field workers at the edges",
     "archaeological, period-authentic, pattern-revealing, multiple"),
    # cue 79 — 553s — "The language does not describe a single anomalous find."
    ("Close-up of the 1873 field notation — plural language clearly legible in the agent's handwriting, the entry explicitly describing multiple individuals, this is not a single measurement error on one specimen",
     "textual, forensic, pattern-revealing, archival"),
    # cue 80 — 560s — VIDEO — "It describes a pattern of oversized remains from a single site."
    ("Push into the 1873 West Virginia field report — camera moves toward the specific passage describing multiple oversized remains from one location, the word 'pattern' implied by the notation's language coming into sharp focus",
     "pattern-revealing, archival, investigative, revelatory"),
    # cue 81 — 567s — "The Bureau of Ethnology. Illinois site. The agent on record is Cyrus Thomas."
    ("1882 Illinois Bureau of Ethnology field report cover page — date 1882 and Illinois heading clearly visible, Cyrus Thomas's name at the top, the third document in the chronological sequence of Bureau excavation records",
     "archival, documentary, transitional, investigative"),
    # cue 82 — 574s — "Cyrus Thomas who would go on to write the 1894 definitive report concluding there was nothing unusual about the mound building peoples."
    ("Side-by-side of Cyrus Thomas's 1882 Illinois field report and his 1894 final published conclusion — the same man, twelve years apart, the pre-policy observation beside the institutional verdict, the distance between them unresolved",
     "contradictory, biographical, investigative, revelatory"),
    # cue 83 — 581s — "His field report from the Illinois site includes measurements and physical descriptions that are, let us say, difficult to reconcile"
    ("Open 1882 Illinois field report by Cyrus Thomas — the specific pages with anomalous measurement notations visible in his pre-publication handwriting, the observational record before the official conclusion was established",
     "archival, biographical, pre-policy, investigative"),
    # cue 84 — 588s — VIDEO — "with his later public conclusions. Now, Cyrus Thomas. I want to be precise here because he is simultaneously"
    ("Slow pan from Cyrus Thomas's 1882 Illinois field notes to his 1894 published Mound Explorations report — camera moves across the desk from the observational document to the institutional conclusion, the temporal distance made visible",
     "contradictory, biographical, investigative, revelatory"),
    # cue 85 — 595s — "one of the most important figures in this story and one of the most complicated. He was"
    ("Portrait of Cyrus Thomas in period scholarly attire — serious Victorian scientist photographed in institutional setting, the weight of his complicated role in this story visible in the formal gravity of the period portrait",
     "biographical, complex, period-authentic, scholarly"),
    # cue 86 — 602s — "a serious scientist, a trained entomologist who came to archaeology mid-career. The 1894"
    ("Victorian entomologist's scientific collection — period specimen boxes of carefully pinned and labeled insects, taxonomic precision in the handwritten labels, the rigorous scientific discipline Thomas brought from entomology into archaeology",
     "scientific, period-authentic, biographical, rigorous"),
    # cue 87 — 609s — "mound builder report he produced is, in most respects, methodologically sound. It correctly"
    ("Internal pages of the 1894 Mound Explorations report — systematic site documentation in the text, precise measurement tables, excavation grids, the legitimate scholarly apparatus of rigorous scientific methodology",
     "scholarly, rigorous, institutional, legitimate"),
    # cue 88 — 616s — VIDEO — "demolished the lost race mythology that had infected popular imagination for decades."
    ("Slow pan across romanticized Lost Race publications being definitively shelved — speculative Mound Builder mythology being archived beside the 1894 report, the popular imagination corrected by institutional scholarship",
     "corrective, scholarly, period-authentic, resolving"),
    # cue 89 — 623s — "The mounds were built by indigenous peoples. Thomas was right about that. The question"
    ("Authentic Hopewell culture artifacts in museum context — copper ornaments, effigy pipes, stone tools and ceramic vessels, the genuine indigenous material culture that Cyrus Thomas correctly attributed to the mound builders",
     "cultural, indigenous, archaeological, confirmed"),
    # cue 90 — 630s — "is what else he was right about and chose not to include. Because there is a difference"
    ("Cyrus Thomas's two bodies of work on a research desk — the 1882 pre-publication field notes beside the 1894 final report, the gap between what he observed and what he chose to publish made physical",
     "investigative, biographical, omission, revelatory"),
    # cue 91 — 637s — "between concluding that the mounds were built by indigenous peoples, which the evidence supports, and concluding that nothing anomalous was found in those mounds."
    ("Two distinct analytical claims on period paper side by side — 'Mounds built by indigenous peoples: documented' on the left, 'Nothing anomalous found: asserted' on the right — the logical distinction between two separate and unequal claims",
     "analytical, logical, comparative, investigative"),
    # cue 92 — 644s — VIDEO — "Those are two separate claims, and only one of them is fully supported by Thomas's own field documentation."
    ("Push toward Thomas's 1882 Illinois field notes — camera moves toward the specific pages with measurement entries that appear in the pre-publication record but not the published conclusions, the unsupported claim's absence made visible",
     "focused, contradictory, investigative, revelatory"),
    # cue 93 — 651s — "Why does his pre-1894 field record not match his 1894 published conclusions? That is a question"
    ("Comparison table on period paper — Thomas's 1882 field report measurement entries beside the 1894 published data for the same Illinois sites, the discrepancies between pre- and post-policy documentation visible line by line",
     "comparative, investigative, discrepancy, analytical"),
    # cue 94 — 658s — "the mainstream literature has not answered to my satisfaction. Let us shift the angle slightly"
    ("Open academic library with rows of post-1894 archaeological reference texts — the settled consensus literature that built on Thomas's official conclusions, none revisiting the discrepancy between his field observations and his published verdict",
     "institutional, settled, unsatisfying, investigative"),
    # cue 95 — 665s — "Let us shift the angle slightly, because so far we have been looking at what was found. Now let us look at what happened afterward."
    ("Evidence board photographed from a new angle — the investigation surface showing the BEFORE section (what was found) shifting perspective to the AFTER section (what happened to it), the custody question replacing the discovery question",
     "transition, investigative, new angle, ominous"),
    # cue 96 — 672s — VIDEO — "Researcher Richard Duhrst, and I want to be clear that Duhrst is an investigative researcher, not a credentialed academic"
    ("Push toward Dewhurst's 2013 book on a research desk — 'The Ancient Giants Who Ruled America' with documentary appendices open, primary source citations visible, the apparatus of serious investigative research rather than fringe speculation",
     "investigative, documentary, sourced, credentialed"),
    # cue 97 — 679s — "but he is also not a random internet personality. His 2013 book, The Ancient Giants Who Ruled America, is heavily sourced"
    ("Open pages of Dewhurst's book showing primary source documentation — footnotes referencing Smithsonian Annual Reports by year, FOIA document reproductions in the appendix, the evidentiary foundation of legitimate investigative journalism",
     "investigative, documentary, methodical, substantial"),
    # cue 98 — 686s — "and includes documentary appendices. Duhrst obtained, via Freedom of Information Act request, an internal Smithsonian memorandum dated 1968."
    ("FOIA request and official government response packet — the formal Freedom of Information Act paperwork with FOIA stamp and case number, a folded internal document visible inside the opened response envelope",
     "governmental, investigative, revelatory, bureaucratic"),
    # cue 99 — 693s — "Let me read you the relevant language, as quoted in Duhrst's published work. The memo"
    ("Open page of Dewhurst's book with the 1968 memo passage quoted in full — the published version of the suppression language in the investigative journalist's treatment of the primary source, surrounding context provided",
     "documentary, investigative, textual, analytical"),
    # cue 100 — 700s — VIDEO — "references a policy directive to 'remove from public record skeletal material that is, quote, incompatible with accepted migration theory.'"
    ("Push into the 1968 Smithsonian internal memo — camera moves slowly toward the typewritten phrase 'remove from public record,' the period government font sharpening as it fills the frame, bureaucratic language with damning implications",
     "damning, bureaucratic, typewritten, ominous"),
    # cue 101 — 707s — "Remove from public record. Now, there are ways to read that phrase charitably."
    ("Close-up of the 1968 memo with 'remove from public record' isolated in a pool of lamplight — the specific policy directive language in typewriter font, the surrounding bureaucratic text providing cold institutional context",
     "damning, bureaucratic, typewritten, isolated"),
    # cue 102 — 714s — "Institutional archiving decisions happen. Collections get reclassified."
    ("Well-organized institutional archive with proper cataloging — orderly shelves of labeled specimen drawers, clear transfer records attached to each, a well-managed collection operating normally, the charitable explanation made concrete",
     "neutral, institutional, orderly, contextual"),
    # cue 103 — 721s — "Material gets transferred to storage, to Native American tribal authorities under NAGPRA"
    ("Formal specimen transfer documentation scene — labeled archival boxes with official transfer forms attached, institutional exchange of properly documented material, the legitimate and paper-trailed process of repatriation",
     "legitimate, formal, documented, procedural"),
    # cue 104 — 728s — VIDEO — "the Native American Graves Protection and Repatriation Act, passed in 1990, or simply lost in the way that large institutional collections lose things over decades."
    ("Pan across the NAGPRA legislation document — the 1990 congressional text, official US government seal, the formal legal framework that requires notification, documentation, and consent for any repatriation of indigenous remains",
     "legislative, formal, governmental, procedural"),
    # cue 105 — 735s — "Those are all real explanations. They would be more satisfying if the Smithsonian had been consistent in applying them."
    ("Disorganized institutional storage room — boxes stacked unlabeled to the ceiling, uncatalogued material accumulated over decades, genuine institutional chaos, a plausible explanation that would be more convincing for these specific specimens",
     "chaotic, institutional, neglect, contextual"),
    # cue 106 — 742s — "But NAGPRA repatriation is documented. It creates paper trails. Tribal nations receive notifications."
    ("NAGPRA documentation table — formal transfer records, tribal notification letters, chain of custody forms all properly filled out, the documented paper trail that every standard repatriation is required by law to generate",
     "documented, formal, procedural, investigative"),
    # cue 107 — 749s — "There are records. But for the specific remains documented in the 1868, 1873, and 1882 Bureau reports"
    ("Three Bureau field reports with conspicuously blank adjacent spaces — the 1868 Ohio, 1873 West Virginia, and 1882 Illinois documents laid out, each with an empty form where the required NAGPRA transfer record should be filed",
     "conspicuous absence, investigative, revealing, ominous"),
    # cue 108 — 756s — VIDEO — "where are the NAGPRA records? Where are the transfer records? Where is the chain of custody"
    ("Slow searching shot through Smithsonian specimen drawers — camera moves down a corridor of institutional storage looking for Bureau accession numbers from 1868, 1873, and 1882, drawer after drawer with no match",
     "institutional, searching, ominous, investigative"),
    # cue 109 — 763s — "for remains that were on the institutional record, excavated by Bureau agents"
    ("Smithsonian collection accession register — period institutional record showing Bureau-recovered specimens entered with official accession numbers, the documented proof that these remains were received into Smithsonian custody",
     "institutional, archival, documented, investigative"),
    # cue 110 — 770s — "and brought into Smithsonian collections? We have asked that question."
    ("Chain of custody document stopping mid-entry — formal institutional record that ends without continuation, the next custodian line blank, the specimens' trail going cold at the point of institutional receipt",
     "broken, investigative, absence, ominous"),
    # cue 111 — 777s — "Researchers have asked that question through official channels."
    ("Formal research inquiry letter on institutional letterhead — a written request to the Smithsonian for access to specific Bureau-documented specimens from 1868, 1873, 1882, the official channel of legitimate academic inquiry",
     "formal, institutional, investigative, procedural"),
    # cue 112 — 784s — VIDEO — "The answer, when an answer comes at all, is that the material cannot be located."
    ("Pull back from Smithsonian response letter — camera begins tight on 'cannot be located' in typewritten text on official stationery, then pulls back to reveal the full institutional letterhead, the formal bureaucratic non-answer",
     "bureaucratic, damning, ominous, cold"),
    # cue 113 — 791s — "Cannot be located. Here is where it gets uncomfortable. The New York Times, November, 1883."
    ("1883 New York Times front page — the full broadsheet of the period's paper of record, November 1883, not fringe, not tabloid, the mainstream institutional press of its era",
     "mainstream, institutional, period-authentic, authoritative"),
    # cue 114 — 798s — "This is not a fringe publication. This is not a sensationalist tabloid. The New York Times of 1883 covered"
    ("1883 NYT masthead detail — the iconic New York Times banner in period typography, the institutional authority of the publication made visible, the credibility of this source established before its content is revealed",
     "credible, institutional, mainstream, authoritative"),
    # cue 115 — 805s — "the Smithsonian Mound Excavation Program as legitimate scientific news, because it was legitimate scientific news."
    ("1883 New York Times article about Smithsonian mound excavations — science reporting with Bureau of Ethnology credited as the source, the mainstream press covering mainstream institutional science as it happened",
     "mainstream, scientific, institutional, documentary"),
    # cue 116 — 812s — VIDEO — "Their coverage specifically references Bureau findings from Ohio Mound sites, describes physical measurements of recovered remains"
    ("Pan across the 1883 NYT article — camera moves over the specific passage mentioning Ohio Bureau findings, the language of physical measurements visible in the period newsprint, the mainstream documentation of the institutional scientific record",
     "documentary, investigative, revelatory, period-authentic"),
    # cue 117 — 819s — "and attributes the work directly to Bureau of Ethnology agents. A mainstream newspaper covering a mainstream scientific institution"
    ("1883 NYT article with Bureau of Ethnology attribution highlighted — the mainstream press explicitly crediting the institutional scientific apparatus, a mainstream source documenting mainstream institutional science",
     "institutional, mainstream, attribution, authoritative"),
    # cue 118 — 826s — "documenting findings that the mainstream scientific institution now says do not exist. That is not a fringe source contradicting an institution."
    ("1883 NYT article beside a contemporary Smithsonian denial — the historical mainstream press documentation set against the institution's current position, the contradiction between the Smithsonian's own 1883 press record and its present stance",
     "contradiction, institutional, documentary, damning"),
    # cue 119 — 833s — "That is the institution's own contemporary press record, contradicting the institution's current position."
    ("Wide shot of the Smithsonian Institution Building at dusk with a superimposed 1883 NYT headline — the institution's physical presence today beside its own 19th century mainstream press record, the internal contradiction made spatial",
     "institutional, contradictory, atmospheric, damning"),
    # cue 120 — 840s — VIDEO — "And here is the part I keep coming back to, genuinely difficult to explain away"
    ("Slow push toward a lone researcher at his desk at night — single pool of lamplight, the 1868 and 1873 Bureau field reports spread before him, the posture of someone wrestling with evidence that resists every comfortable explanation",
     "contemplative, night, investigative, human, uneasy"),
    # cue 121 — 847s — "The 1968 internal memo does not say we have reexamined these remains and determined the original measurements were inaccurate."
    ("Close-up of the 1968 Smithsonian internal memo — typewritten text visible, conspicuously no language reexamining measurements, no scientific review cited, no laboratory analysis — the memo is not a scientific document",
     "absence, damning, typewritten, forensic"),
    # cue 122 — 854s — "It does not say these specimens have been repatriated under tribal agreement."
    ("The 1968 memo viewed full frame — NAGPRA language conspicuously absent, no tribal agreement referenced anywhere in the typewritten text, not a repatriation directive",
     "absence, ominous, institutional, forensic"),
    # cue 123 — 861s — "It does not provide a scientific correction."
    ("The full 1968 Smithsonian memo under direct light — no laboratory analysis, no revised measurements, no peer-reviewed correction visible in any line, the complete absence of scientific methodology in a supposedly institutional document",
     "absence, damning, cold, institutional"),
    # cue 124 — 868s — VIDEO — "It says remove from public record, material incompatible with accepted migration theory"
    ("Slow push into the key phrase of the 1968 memo — 'remove from public record' growing in the frame, typewritten characters sharp against aged paper, the most consequential institutional directive in the documentary record",
     "damning, revelatory, focused, typewritten"),
    # cue 125 — 875s — "incompatible with accepted migration theory, not incompatible with the physical evidence"
    ("Two-column comparison laid on a document desk — left column: 'Incompatible with migration theory'; right column: 'Incompatible with physical evidence' — the critical distinction between what the memo says and what a scientific correction would say",
     "analytical, distinction, revelatory, critical"),
    # cue 126 — 882s — "not incompatible with biological analysis, incompatible with a theory"
    ("Extreme close-up of the phrase 'incompatible with a theory' in the 1968 memo typeface — the word 'theory' isolated and in sharp focus, the admission that suppression was theoretical not empirical",
     "focused, damning, typewritten, revelatory"),
    # cue 127 — 889s — "a theoretical framework about when and how humans arrived in the Americas. That is a very different thing."
    ("Aged cartographic Bering Land Bridge map — the theoretical migration framework being referenced by the 1968 memo, Beringia labeled with migration arrows, the model that was treated as more important than physical evidence",
     "theoretical, geographic, cartographic, contextual"),
    # cue 128 — 896s — VIDEO — "That is a very different thing. And I think that difference matters."
    ("Slow pull-back from close-up of the 1968 memo to reveal the full documentary landscape — Bureau field reports, chain-of-custody gaps, FOIA packet surrounding the memo, the difference between theory and evidence made spatial",
     "widening, contextual, revelatory, investigative"),
    # cue 129 — 903s — "The Bering Land Bridge model, the accepted framework for human migration into the Americas"
    ("Aged parchment map of the Bering Land Bridge — the foundational migration framework, Beringia land connection between northeastern Asia and North America, migration arrows drawn across in period cartographic style",
     "theoretical, geographic, foundational, cartographic"),
    # cue 130 — 910s — "holds that the first peoples crossed from northeastern Asia into North America"
    ("Large-format period map showing the northeastern Asia to Alaska migration route — the Beringia crossing emphasized, the foundational geographic claim of American prehistory made visual and concrete",
     "geographic, foundational, cartographic, definitive"),
    # cue 131 — 917s — "via the Beringia land connection, roughly 15,000 to 20,000 years ago"
    ("Artist's reconstruction of the Beringia land mass 20,000 years ago — ancient tundra landscape connecting Asia to North America, migration figures moving across the land bridge in period archaeological illustration style",
     "reconstructed, ancient, prehistoric, geographic"),
    # cue 132 — 924s — VIDEO — "That model has been revised repeatedly. The Clovis first hypothesis, around 13,000 years ago"
    ("Pan across Clovis culture spear points in museum display — the distinctive fluted stone tools dated 13,000 BP, the hypothesis that placed first migration at a fixed date now being challenged by new sites",
     "archaeological, foundational, institutional, revisable"),
    # cue 133 — 931s — "has been substantially challenged by sites like Monteverde in Chile"
    ("Monteverde archaeological site — southern Chile, ancient human artifacts in stratigraphic context, the location that pushed the American arrival date back before the Clovis threshold, the primary challenger site",
     "challenging, southern Chile, archaeological, ancient"),
    # cue 134 — 938s — "Meadowcroft Rock Shelter in Pennsylvania, and the Chiqui Huité cave findings in Mexico"
    ("Deep rock shelter cave excavation — ancient human artifacts visible in geological stratigraphy, the confined excavation space containing evidence that challenges the 13,000-year ceiling of the standard model",
     "cave, ancient, archaeological, challenging"),
    # cue 135 — 945s — "which push confirmed human presence in the Americas back considerably further. The model is not static."
    ("Migration timeline revision diagram — the Clovis first date being pushed progressively earlier by new site discoveries, the legitimate scholarly revision of the mainstream model shown across multiple decades",
     "evolving, scholarly, timeline, revision"),
    # cue 136 — 952s — VIDEO — "The mainstream will tell you that. And they are right. Migration theory has evolved over 30 years."
    ("Pan across contemporary peer-reviewed archaeological journals — publications from the 1990s through 2020s tracking the evolution of migration theory, the honest acknowledgment that the mainstream model has changed",
     "scholarly, contemporary, legitimate, evolving"),
    # cue 137 — 959s — "But here is what it has not done. It has not made room for a population of unusually large-statured humans"
    ("Bering Land Bridge standard migration model — the official depiction of the arriving populations, uniform in stature, the conspicuous absence of biological variation the model has refused to accommodate",
     "absence, constrained, institutional, ominous"),
    # cue 138 — 966s — "in the pre-Columbian North American interior. Not because the skeletal evidence rules that out"
    ("Vast pre-Columbian North American interior landscape — Ohio valley, burial mounds as low grassy rises in the autumn distance, the territory where the skeletal evidence was found and then lost",
     "ancient, geographic, vast, Ohio"),
    # cue 139 — 973s — "but because the skeletal evidence for it has, by a remarkable coincidence, become largely inaccessible"
    ("Heavy padlock on an institutional filing cabinet labeled with Bureau of Ethnology accession numbers — RESTRICTED tag attached, key conspicuously absent, a row of identical sealed cabinets behind",
     "locked, inaccessible, ominous, conspiratorial"),
    # cue 140 — 980s — VIDEO — "Ross Hamilton, an Ohio-based researcher who has spent decades working with mound site documentation in the Hopewell Heartland"
    ("Hamilton's research desk — Ohio Hopewell territory map covered in annotated site markers, Bureau field reports cross-referenced, red thread connecting burial mound locations across the state map",
     "investigative, Ohio, cross-referenced, Hamilton"),
    # cue 141 — 987s — "has corroborated the field record pattern we have been discussing"
    ("Hamilton's cross-referenced documentation — multiple Bureau field reports aligned with site location maps, the consistent pattern emerging across the documentary evidence from three separate states",
     "cross-referenced, corroborating, investigative, pattern"),
    # cue 142 — 994s — "Hamilton's work, specifically his extensive cross-referencing of field agent reports with site location records"
    ("Field agent reports laid side-by-side with period site location maps — the specific cross-referencing methodology, connection lines drawn between measurement notations and their geographic excavation sites",
     "methodical, cross-referenced, specific, investigative"),
    # cue 143 — 1001s — "suggests that the oversized remains were not random anomalies scattered across excavation sites"
    ("Non-random distribution map — burial mound sites across the Ohio interior with anomalous find locations marked in red, the clustering clearly non-random, the geographic specificity of the pattern unmistakable",
     "non-random, clustering, geographic, investigative"),
    # cue 144 — 1008s — VIDEO — "They appear disproportionately in high-status burial contexts. Mounds with exceptional artifact density."
    ("Push into a high-status Hopewell burial context — camera approaches an excavation with exceptional artifact density, copper ornaments and shell beads arranged around skeletal remains, the elite interment markers unmistakable",
     "elite, high-status, archaeological, revelatory"),
    # cue 145 — 1015s — "Mounds associated with what archaeologists recognize as elite or ceremonial interment"
    ("Elite Hopewell burial context in active excavation — copper headdress fragments, shell bead necklace, ceremonial objects in dark ochre soil, the exceptional density of grave goods marking this as the highest social tier",
     "elite, ceremonial, archaeological, ancient"),
    # cue 146 — 1022s — "What does it mean if the largest individuals were consistently buried in the highest status contexts?"
    ("Archaeological status correlation diagram — burial contexts ranked by artifact density, with exceptional skeletal stature consistently occupying the highest-status positions across multiple independent sites",
     "analytical, correlation, pattern-revealing, investigative"),
    # cue 147 — 1029s — "It suggests they were not outliers to the culture. They were central to it."
    ("Overhead view of Hopewell burial mound interior — the primary central burial at the focal point, exceptional in size and artifact density, secondary burials surrounding it at the margins, centrality made spatial",
     "central, overhead, archaeological, significant"),
    # cue 148 — 1036s — VIDEO — "They held specific social roles — leadership, priesthood — something we do not have a clean, modern category for"
    ("Push into a ceremonial Hopewell burial reconstruction — elaborate grave goods arranged around the central interment, copper headdress, ceremonial regalia, the elevated social role made visible through the ritual treatment",
     "ceremonial, ancient, elite, mysterious"),
    # cue 149 — 1043s — "reflected in how they were buried. That is a coherent interpretation of the field record."
    ("Hamilton's multi-site field correlation — Bureau reports from Ohio, West Virginia, and Illinois each showing the same stature-to-status pattern, the coherent interpretation emerging across independent documentary sources",
     "multi-site, coherent, pattern, investigative"),
    # cue 150 — 1050s — "It significantly complicates the standard model of Hopewell and Adena social organization"
    ("Standard Hopewell social organization diagram annotated with conflicting evidence — the institutional model with Hamilton's anomalous pattern data points overlaid, the clean official structure visibly complicated",
     "complicating, analytical, conflict, institutional"),
    # cue 151 — 1057s — "Look, I know how this sounds. I know that the moment you say giant skeletons"
    ("Sensationalist tabloid media wall — lurid giant skeleton headlines, exaggerated engravings, the cultural noise that immediately arises with the topic and contaminates legitimate inquiry",
     "sensationalist, cultural noise, contrast, honest"),
    # cue 152 — 1064s — VIDEO — "you lose a significant portion of your audience to associations with ancient aliens"
    ("Pan across sensationalist ancient aliens style media — cable TV thumbnails, internet show stills, the category of content that has poisoned public perception of the legitimate primary source question",
     "sensationalist, contrast, cultural, corrupted"),
    # cue 153 — 1071s — "every bad faith use of this material that has been made over the past 30 years. I am asking you to set that association aside."
    ("Researcher's hand pushing aside a stack of sensationalist giant skeleton publications — clearing a space on the desk, the act of deliberately setting the mythology aside to access the primary source documents underneath",
     "dismissive, focused, transition, investigative"),
    # cue 154 — 1078s — "Not because the sensationalist versions of this story are credible. They mostly are not."
    ("Pile of discredited sensationalist publications — fringe books and cable TV show promotional materials honestly portrayed as the low-credibility content they are, set apart from the sober Bureau documents beside them",
     "honest, discredited, contextual, clear-eyed"),
    # cue 155 — 1085s — "But because the core documentary question is separate from the mythology built around it. The core question is this."
    ("Clean research desk with only primary source documents — just the 1868, 1873, and 1882 Bureau field reports under a single lamp, mythology stripped away, the documentary core of the question isolated",
     "focused, documentary, clean, investigative"),
    # cue 156 — 1092s — VIDEO — "Did Bureau of Ethnology field agents document skeletal remains of exceptional stature in their official reports?"
    ("Pan across Bureau of Ethnology field reports — 1868, 1873, 1882 documents with measurement entries visible, the official institutional record being examined to answer the first and most fundamental documentary question",
     "documentary, investigative, primary source, factual"),
    # cue 157 — 1099s — "The answer, based on the primary source record, appears to be yes."
    ("Three Bureau reports with measurement notations circled — the 1868 Ohio, 1873 West Virginia, and 1882 Illinois field reports each showing an exceptional stature entry circled in red, the affirmative answer across three separate documents",
     "affirmative, multi-source, documentary, revelatory"),
    # cue 158 — 1106s — "Did those remains subsequently become inaccessible in a manner inconsistent with normal archival or repatriation procedures?"
    ("Empty institutional specimen drawers labeled for Bureau-recovered specimens — accession numbers on the tabs, drawers completely empty inside, no repatriation records attached, no transfer documentation visible",
     "inaccessible, absence, institutional, ominous"),
    # cue 159 — 1113s — "The evidence suggests yes. Does an internal memo from 1968 reference a policy of removing anomalous material from the public record?"
    ("The 1968 Smithsonian internal memo beside empty accession drawers — the suppression directive and the evidence of its implementation in one frame, the second and third questions answered by the same documentary pairing",
     "damning, connecting, documentary, ominous"),
    # cue 160 — 1120s — VIDEO — "Richard Dewhurst's FOIA documentation says yes. I am not saying this is proof of a global conspiracy."
    ("Push toward Dewhurst's FOIA response packet on a methodical research desk — the documentary mechanism of access, primary sources organized beside it, the responsible framing that separates documentary evidence from conspiracy framing",
     "confirmatory, measured, documentary, responsible"),
    # cue 161 — 1127s — "I am not saying the Smithsonian is hiding a race of giants."
    ("Calm methodical research setting — primary sources organized carefully on a clean desk, no sensationalist elements, no conspiratorial imagery, just documents and a researcher's careful attention to what they actually show",
     "measured, responsible, methodical, honest"),
    # cue 162 — 1134s — "I am saying that the institution's own paper trail points toward a deliberate decision"
    ("Smithsonian's own documentary sequence laid chronologically — Annual Reports, 1968 internal memo, FOIA response laid in order, the institution's own documents making the institutional argument against itself",
     "self-indicting, institutional, paper trail, documentary"),
    # cue 163 — 1141s — "to suppress specific categories of physical evidence, and that the justification given for that suppression"
    ("The suppression paper trail laid out in full arc — Bureau field finds, institutional custody records, 1968 suppression memo, empty accession drawers — each document in the chain on the same desk",
     "chronological, suppression, documentary, ominous"),
    # cue 164 — 1148s — VIDEO — "was theoretical consistency, not scientific accuracy. Those are meaningfully different things"
    ("Two-column visual comparison — 'Theoretical Consistency' on the left with migration model diagram, 'Scientific Accuracy' on the right with physical measurement data — camera moves between them to show the distinction",
     "analytical, distinction, critical, documentary"),
    # cue 165 — 1155s — "and in the practice of science, they should be treated as meaningfully different things. Louis Henry Morgan"
    ("Period scientific methodology context giving way to a portrait — the principled epistemological distinction the narrator is drawing, transitioning toward the pre-institutional observer who illustrates it",
     "methodological, scientific, principled, transitional"),
    # cue 166 — 1162s — "the 19th century ethnologist, one of the foundational figures of American anthropology"
    ("Lewis Henry Morgan at a field documentation desk — Victorian-era ethnologist in period dress, a large-format field ledger open before him, the foundational American anthropologist at candid pre-institutional scholarly work",
     "biographical, period-authentic, scholarly, candid"),
    # cue 167 — 1169s — "documented mound builder social structures and physical evidence before the suppression era began in earnest."
    ("Morgan's pre-institutional field documentation — his 1850s and 1860s field notes on mound builder populations, the candid physical observations made before any institutional consensus existed to constrain what could be recorded",
     "pre-institutional, candid, biographical, documentary"),
    # cue 168 — 1176s — VIDEO — "Morgan's work is interesting precisely because he was working before there was an institutional consensus to protect."
    ("Pan across Morgan's pre-Bureau field notes — camera moves across the period handwriting of his systematic pre-consensus observations, the scholarly freedom visible in what he felt free to observe and record",
     "pre-consensus, candid, scholarly, free"),
    # cue 169 — 1183s — "His observations about the physical diversity of mound builder remains, made in the 1850s and early 1860s, are more candid"
    ("Open page from Morgan's 1850s-1860s ethnological field notes — specific passages noting significant physical variation in mound builder populations, 1850s dating visible, the pre-Bureau candid observational record",
     "candid, specific, archival, pre-institutional"),
    # cue 170 — 1190s — "than the later Bureau literature. Not because he was a better scientist — the later Bureau work was in many ways more rigorous"
    ("Morgan's candid 1855 field notes beside the more rigorous 1882 Bureau field reports — different eras of scholarship, the later work methodologically superior but the earlier more candid about physical variation",
     "comparative, candid versus constrained, investigative"),
    # cue 171 — 1197s — "but because he was not yet operating under the pressure of theoretical consistency. What Morgan's pre-Bureau documentation suggests"
    ("Morgan at fieldwork in the 1850s — the context of intellectual freedom before the migration framework became institutional policy, a scholar observing without a theoretical conclusion he needed to protect",
     "free, scholarly, pre-institutional, candid"),
    # cue 172 — 1204s — VIDEO — "is that the mound builder populations were not physically homogenous, that there was significant variation"
    ("Pan across diverse skeletal illustrations from Morgan's era — documented physical variation in mound builder populations, the heterogeneity visible in the pre-synthesis observational record before it was smoothed away",
     "diversity, pre-institutional, documentary, candid"),
    # cue 173 — 1211s — "including variation at the upper range of human stature, that the later official synthesis smoothed away now."
    ("Period scientific illustration of stature range — a comparative diagram showing the documented range of human height among mound builder populations, the upper end extending beyond the ordinary, the data the later synthesis erased",
     "scientific, comparative, range, period-authentic"),
    # cue 174 — 1218s — "What would a nine-foot skeleton, or even a consistent population of individuals in the seven to eight foot range"
    ("Human scale comparison illustration — the proportional relationship between standard human height and the seven-to-nine foot range documented by Bureau field agents, the physical scale of the question made visual",
     "scale, confrontational, forensic, visual"),
    # cue 175 — 1225s — "do to the Bering Land Bridge narrative? That depends entirely on when those individuals lived."
    ("Bering Land Bridge map annotated with a question mark — the standard migration model confronted with the timing question, the answer depending entirely on the temporal position of the anomalous population",
     "questioning, theoretical, geographic, destabilizing"),
    # cue 176 — 1232s — VIDEO — "If they were part of the Hopewell or Adena cultures, roughly 500 BCE to 500 CE"
    ("Pan across Hopewell and Adena cultural timeline — the 500 BCE to 500 CE window highlighted on a period cultural map, the specific era being placed before the biological challenge is articulated",
     "temporal, historical, contextual, analytical"),
    # cue 177 — 1239s — "then they do not necessarily challenge the Bering Migration model on timing. They challenge it on biology."
    ("Migration timing diagram pivoting to biological profile — the Bering model intact on the temporal axis, with the biological characteristics axis showing the divergence, the specific and unusual nature of the challenge",
     "analytical, pivoting, investigative, distinction"),
    # cue 178 — 1246s — "A population of consistently large-statured individuals in the North American interior"
    ("North American interior map with population anomaly marker — the Ohio-Indiana-Illinois-Kentucky territory highlighted as the location of the documented anomalous population, the geographic concentration of the evidence",
     "geographic, specific, anomalous, investigative"),
    # cue 179 — 1253s — "with no clear counterpart in the Asian populations they supposedly descended from"
    ("Comparative anthropological illustration — period scientific diagram comparing North American mound builder physical characteristics with Northeast Asian populations, the biological gap between the supposed ancestors and descendants visible",
     "comparative, biological, gap, investigative"),
    # cue 180 — 1260s — VIDEO — "raises questions about parallel or earlier migration routes — coastal routes, Pacific routes"
    ("Pan across alternative migration route map — Pacific coastal and trans-oceanic routes marked alongside the Bering dominant model, the geographic diversity of possible human arrival routes the dominant framework has resisted",
     "alternative, geographic, investigative, open-ended"),
    # cue 181 — 1267s — "routes that the Bering dominant model has historically been slow to accommodate."
    ("Academic consensus diagram showing Bering dominance — the theoretical framework's historic resistance to incorporating alternative routes, the rigidity of the dominant model in the face of complicating evidence",
     "alternative, investigative, geographic, resistant"),
    # cue 182 — 1274s — "And if they predate the Hopewell — if the high-status burial contexts containing oversized remains belong to an earlier cultural layer"
    ("Archaeological stratigraphy cross-section of a burial mound — visible soil layers in excavation trench wall, Hopewell-era material clearly above, an earlier cultural layer with different material characteristics below",
     "stratigraphic, archaeological, earlier, layered"),
    # cue 183 — 1281s — "one that the Hopewell inherited rather than created"
    ("Pre-Hopewell cultural layer exposed in excavation — older period artifacts and soil characteristics visible beneath Hopewell-era deposits, the foundation the later culture built on rather than originated",
     "older, inherited, archaeological, stratigraphic"),
    # cue 184 — 1288s — VIDEO — "then you are potentially looking at a founding population whose biological profile does not match the current migration narrative at all."
    ("Push toward pre-Hopewell burial stratum in excavation — camera moves toward the earlier archaeological layer beneath the known Hopewell context, the founding population whose biological profile is the unresolved question",
     "anomaly, biological, geographic, investigative"),
    # cue 185 — 1295s — "That is not a comfortable question. It is also not one you are allowed to ask at a mainstream archaeological conference"
    ("Formal academic archaeological conference setting — podium and audience in institutional setting, formally dressed academics, the professional environment where certain questions are implicitly forbidden by career consequences",
     "institutional, forbidden, academic, ominous"),
    # cue 186 — 1302s — "without ending your academic career fairly quickly."
    ("Academic career institutional documents — CV, tenure review materials, university appointment letter, the professional infrastructure whose continuation depends on adherence to the consensus theoretical framework",
     "institutional, professional, career, ominous"),
    # cue 187 — 1309s — "Which is, I would argue, itself a piece of evidence worth considering."
    ("Evidence board with career suppression added as a datum — the academic consequence of this line of inquiry treated as its own piece of evidence, 'asking about pre-Hopewell anomalous stature = professionally ended' as a self-referential data point",
     "institutional pressure, suppression, academic, investigative"),
    # cue 188 — 1316s — VIDEO — "Richard Dewhurst's work, whatever its limitations as investigative journalism rather than peer-reviewed scholarship"
    ("Push toward Dewhurst's open book on a research desk — the investigative journalism work with its extensive citations and primary source appendices, its nature as journalism rather than peer-reviewed scholarship honestly acknowledged",
     "honest, comparative, documentary, investigative"),
    # cue 189 — 1323s — "identified over 1,500 newspaper accounts of giant skeleton discoveries in the 19th century American press"
    ("Wall covered with 19th century newspaper clippings about giant skeleton finds — the scale of 1,500 accounts visualized in the sheer volume of accumulated mainstream press documentation from across the country",
     "volume, accumulation, journalistic, overwhelming"),
    # cue 190 — 1330s — "cross-referenced with known excavation site locations. The geographic clustering of those reports is not random."
    ("Cross-reference map — 19th century American newspaper accounts plotted as location dots on a map of mound builder territory, the geographic correlation between press reports and known Bureau excavation sites clearly visible",
     "geographic, cross-referenced, correlating, investigative"),
    # cue 191 — 1337s — "It tracks the mound builder cultural territory with a specificity that makes the all-hoaxes dismissal difficult to sustain."
    ("Geographic clustering map — newspaper account locations forming distinct clusters over Hopewell and Adena territory, the same territory as the Bureau excavations, the statistical improbability of the all-hoaxes explanation made spatial",
     "clustering, geographic, non-random, pattern-revealing"),
    # cue 192 — 1344s — VIDEO — "All of them hoaxes? Every single one? In Ohio and Illinois and West Virginia and Kentucky?"
    ("Slow pan across a four-state map — Ohio, Illinois, West Virginia, Kentucky highlighted in sequence, Bureau excavation sites marked across all four states, the geographic scale the all-hoaxes explanation must account for",
     "geographic, scale, challenging, rhetorical"),
    # cue 193 — 1351s — "By field agents working under institutional supervision"
    ("Bureau of Ethnology institutional hierarchy — chain of command from Smithsonian leadership through Bureau director to supervised field agents in the territories, the accountability structure making unsupervised fabrication organizationally implausible",
     "institutional, hierarchical, supervised, documentary"),
    # cue 194 — 1358s — "submitting reports to the Bureau of Ethnology in the same document series used to record routine and uncontroversial findings."
    ("Bureau Annual Report spread showing routine and anomalous entries in identical institutional format — weather observations, agricultural surveys, routine findings, and exceptional skeletal measurements all formatted identically",
     "contextual, institutional, same series, investigative"),
    # cue 195 — 1365s — "That is a lot of coordinated fabrication for an institution that was, at the time, genuinely trying to produce serious science."
    ("Bureau of Ethnology's serious scientific apparatus — systematic excavation grids, precise measurement protocols, cross-referenced field report system, the institutional infrastructure that makes coordinated fabrication at scale implausible",
     "institutional, rigorous, implausibility, investigative"),
    # cue 196 — 1372s — VIDEO — "Let us bring this together. Not with a conclusion. I do not think the evidence yet supports a definitive conclusion."
    ("Pull back from the full investigation board — camera slowly retreats from the complete primary source surface showing all documents in order, the full picture of the documentary case becoming visible as it recedes",
     "summarizing, documentary, pulling back, deliberate"),
    # cue 197 — 1379s — "But with a clear statement of where the record actually stands."
    ("Complete documentary record arranged chronologically — 1868, 1873, 1882 Bureau field reports and the 1968 FOIA memo laid out in sequence, the full extent of what is actually documented stated without a conclusion imposed",
     "honest, open-ended, chronological, documentary"),
    # cue 198 — 1386s — "Between 1868 and 1889, the Smithsonian Bureau of Ethnology produced field reports documenting skeletal remains of exceptional stature"
    ("Stack of Bureau field reports from 1868-1889 — the specific 21-year window of primary documentation, each report clearly dated, the accumulation of the pre-settlement observational record across Ohio, West Virginia, and Illinois",
     "chronological, accumulating, primary source, documentary"),
    # cue 199 — 1393s — "from burial mound sites across Ohio, West Virginia, and Illinois. Those reports were submitted under the names of credentialed bureau agents."
    ("Three-state documentation map with credentialed agent names — Ohio, West Virginia, Illinois highlighted, Bureau excavation sites marked, field agents' names visible on the document covers across the geographic territory",
     "geographic, specific, credentialed, documentary"),
    # cue 200 — 1400s — VIDEO — "They appeared in official Smithsonian annual report series. They were covered in real time by the mainstream press."
    ("Pan across Bureau Annual Reports beside 1883 New York Times coverage — official institutional documents with annual report designation beside the mainstream press record, institutional and journalistic corroboration in the same frame",
     "corroborating, institutional, mainstream, documentary"),
    # cue 201 — 1407s — "None of those remains are currently accessible in Smithsonian collections."
    ("Empty institutional specimen drawers for Bureau-recovered specimens — labeled tabs for Ohio, West Virginia, and Illinois Bureau excavations, all drawers empty inside, the absence confirmed by the institutional record",
     "absence, ominous, institutional, confirmed"),
    # cue 202 — 1414s — "An internal memo obtained via FOIA request and published by Richard Dewhurst in 2013"
    ("Dewhurst's FOIA response packet and 2013 book together — the formal government FOIA response documents alongside his publication, the mechanism of access and the published presentation of the primary source evidence",
     "governmental, procedural, FOIA, documentary"),
    # cue 203 — 1421s — "references a policy directive to remove from public record any skeletal material incompatible with accepted migration theory."
    ("The 1968 memo key passage isolated — 'remove from public record' and 'incompatible with accepted migration theory' in typewritten text, the complete policy directive language in sharp focus",
     "damning, focused, typewritten, ominous"),
    # cue 204 — 1428s — VIDEO — "The justification given is theoretical, not scientific."
    ("Push into 'incompatible with accepted migration theory' in the 1968 memo — the word 'theory' coming into sharp focus, the theoretical rather than scientific basis of the suppression directive isolated in the frame",
     "focused, damning, distinction, revelatory"),
    # cue 205 — 1435s — "The mainstream explanation — hoax journalism, measurement error, and normal archival attrition — does not account for the full pattern."
    ("Three mainstream explanations as column headers on a document — Hoax Journalism, Measurement Error, Archival Attrition — each with the evidence beneath it that the explanation fails to cover, falling short of the full pattern",
     "analytical, comparative, insufficient, investigative"),
    # cue 206 — 1442s — "It does not explain the institutional memo."
    ("The 1968 memo alone on a clean desk — the document that cannot be explained by measurement error or archival neglect, the suppression directive that stands outside every charitable explanation that has been offered",
     "unexplained, singular, damning, ominous"),
    # cue 207 — 1449s — "It does not explain the absence of chain of custody records for documented excavations."
    ("Bureau accession register showing documented specimens beside a blank chain of custody form — institutional receipt confirmed on one side, the custody trail that ends without explanation on the other",
     "unexplained, documentary, absence, investigative"),
    # cue 208 — 1456s — VIDEO — "It does not explain why the anomalies cluster in high-status burial contexts rather than appearing randomly across excavation sites."
    ("Distribution map showing non-random clustering — burial sites with anomalous stature finds consistently at high-status mound positions, the non-random geographic pattern that random measurement error cannot explain",
     "non-random, clustering, geographic, unresolved"),
    # cue 209 — 1463s — "What does a nine-foot skeleton do to the Bering Land Bridge narrative? I think the question answers itself."
    ("Bering Land Bridge migration map confronted with the stature question — the standard migration diagram face-to-face with the scale of the anomaly, the question that answers itself by the act of being clearly asked",
     "rhetorical, confrontational, geographic, revelatory"),
    # cue 210 — 1470s — "It forces us to ask who else was here and when and where they came from."
    ("Pre-Columbian North America map with three question marks — 'who?', 'when?', and 'where from?' placed over the continental interior, the three unanswered questions made spatial across the territory of the evidence",
     "open, questioning, geographic, investigative"),
    # cue 211 — 1477s — "And it forces us to ask whether the theoretical framework we have been given for pre-Columbian North America is a scientific conclusion or a managed story."
    ("The Bering Land Bridge model diagram beside an investigative evidence board — the official theoretical framework confronted with the 'managed story' question, the core distinction the entire documentary record has been building toward",
     "climactic, investigative, rhetorical, revelatory"),
    # cue 212 — 1484s — VIDEO — "The question is not whether these remains existed. The field record says they did. The question is why we are not allowed to ask what happened to them."
    ("Pull back from Bureau field reports to reveal the full documentary record — the three primary source documents in frame confirming existence, pulling back to the unanswered question of what happened afterward",
     "climactic, investigative, documentary, unanswered"),
    # cue 213 — 1491s — "That is the question that no official source has answered."
    ("Silhouetted researcher standing in an empty government archive room — rows of dark filing cabinets stretching into shadow, one drawer hanging open and empty, a single desk lamp the only light, the weight of the unanswered question made physical",
     "haunting, institutional, unanswered, shadow"),
    # cue 214 — 1498s — outro bridge — investigation dissolves to landscape
    ("Slow dissolve from archival papers on a desk to an exterior dusk sky over an Ohio river valley — the documentary world giving way to the landscape where the evidence was found, contemplative and atmospheric",
     "transitional, reflective, dissolving, atmospheric"),
    # cue 215 — 1505s — "If this is the first Ruins Untold video you have watched, welcome. This channel exists for exactly this kind of question."
    ("Wide golden-hour shot of the Newark Great Circle earthwork — late afternoon light raking low across the ancient grass geometry, the massive circular form rolling across the Ohio landscape, welcoming and vast",
     "welcoming, golden hour, earthwork, channel identity"),
    # cue 216 — 1512s — VIDEO — "The gap between what the physical evidence shows and what the institutional narrative acknowledges. We follow the evidence, even when it leads"
    ("Slow pan across the face of the Smithsonian Institution Building in Washington D.C. at dusk — the neoclassical sandstone facade lit from below, institutional authority made monumental against a darkening sky",
     "institutional, dusk, authority, contrast"),
    # cue 217 — 1519s — "somewhere uncomfortable. If you have been here before, you know the drill."
    ("Wide atmospheric shot of an Ohio grass burial mound at twilight — the mound silhouetted against a pink and orange horizon, ancient and solitary in the open landscape, the site of the evidence and the unanswered question",
     "atmospheric, twilight, burial mound, ancient"),
    # cue 218 — 1526s — "Have you encountered the Bureau of Ethnology reports directly? Have you tried to access any of this material through official channels?"
    ("Close-up of hands on a desk with a partially filled FOIA request form beside a copy of the 1868 Smithsonian Annual Report — the investigator engaged in the formal act of seeking access, the community of inquiry made personal",
     "engaged, investigative, community, FOIA"),
    # cue 219 — 1533s — "Tell me what you found. And if you have not already, subscribe."
    ("Wide aerial shot of the Newark Earthworks complex at golden hour — the full geometry of the ancient octagon and circle visible from above, the invitation to look deeper at what has been here all along",
     "aerial, golden hour, invitation, earthworks"),
    # cue 220 — 1540s — VIDEO — "Not because the algorithm rewards it, though it does, but because the next video in this series goes somewhere I have been building toward for a long time."
    ("Slow drone pullback from a burial mound at dusk — the earthwork shrinking as the camera lifts, the Ohio landscape opening up around it, the river valley and the horizon emerging, anticipation of what comes next",
     "aerial, pullback, anticipatory, dusk"),
    # cue 221 — 1547s — "We are going to look at what the mounds themselves are built on top of. The geological layer beneath the earthworks."
    ("Geological cross-section illustration of a Hopewell burial mound in profile — layers of earthen fill clearly labeled above a distinct older geological stratum, the scientific diagram revealing what exists beneath the visible surface",
     "geological, cross-section, revealing, scientific"),
    # cue 222 — 1554s — "The geological layer beneath the earthworks."
    ("Extreme close-up of a soil core sample extracted from beneath a burial mound — distinct dark pre-Hopewell strata visible in the cylinder, the ancient layer that predates the builders, the physical evidence of an older presence",
     "geological, close-up, soil core, ancient"),
    # cue 223 — 1561s — "And what the dating of that layer suggests about who was here, before the Hopewell. Come back for that one. I think you will want to see it."
    ("Silhouette of a lone figure standing at the top of a grass burial mound at twilight — looking out over the river valley below, the vast darkening landscape ahead, the invitation to return for what the next layer reveals",
     "silhouetted, twilight, forward-looking, invitation"),
]

SHOT_ROTATION   = ["extreme close-up", "medium shot", "wide establishing shot",
                   "close-up", "overhead looking down", "low-angle looking up",
                   "medium shot", "close-up"]
MOTION_ROTATION = ["slow dolly push-in", "low tracking shot", "gentle tilt-up reveal",
                   "slow pan across scene", "aerial pull-back",
                   "push through archival shelves"]

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Generate visual cue schedule
# ─────────────────────────────────────────────────────────────────────────────
image_prompts = []
video_prompts = []
pinned_used   = []
drift_entries = []

img_idx = vid_idx = cue_n = 0
cur = 0.0

while cur < ACTUAL_DUR - 3:
    ts  = sec_to_ts(cur)
    n_t = nearest_text(max(0.0, cur - J_CUT_OFFSET))

    # ── Pinned ─────────────────────────────────────────────────────────────
    pinned = check_pinned(n_t)
    if pinned and pinned["asset_type"] == "video":
        dur      = float(pinned["duration_seconds"])
        filename = f"video_{vid_idx+1:04d}.mp4"
        video_prompts.append({
            "index": vid_idx+1, "timestamp": ts,
            "duration_seconds": dur, "filename": filename,
            "script_line": n_t[:80],
            "pinned": True, "pinned_id": pinned["id"],
            "pinned_source": pinned["asset_path"],
            "copy_instruction": f"Copy {pinned['asset_path']} → {PROJ}/videos/{filename}",
            "veo_prompt": "[PINNED CHANNEL ASSET — do not generate, use source file]",
            "nb2_prompt": {
                "shot_type": "channel intro", "subject": pinned["description"],
                "action": "pinned asset", "environment": "channel intro",
                "style": STYLE, "color_palette": PALETTE, "lighting": LIGHTING,
                "mood": "channel intro, cosmic, mysterious",
                "negative_constraints": NEG_VID,
            }
        })
        pinned_used.append({"id": pinned["id"], "ts": ts, "filename": filename})
        vid_idx += 1; cue_n += 1
        cur += IMAGE_INTERVAL
        continue

    elif pinned and pinned["asset_type"] == "image":
        filename = f"image_{img_idx+1:04d}.png"
        image_prompts.append({
            "index": img_idx+1, "timestamp": ts,
            "duration": IMAGE_INTERVAL, "filename": filename,
            "script_line": n_t[:80],
            "pinned": True, "pinned_id": pinned["id"],
            "pinned_source": pinned["asset_path"],
            "copy_instruction": f"Copy {pinned['asset_path']} → {PROJ}/images/{filename}",
            "nb2_prompt": {
                "shot_type": "channel asset", "subject": pinned["description"],
                "action": "pinned asset", "environment": "channel intro",
                "style": STYLE, "color_palette": PALETTE, "lighting": LIGHTING,
                "mood": "channel intro", "negative_constraints": NEG,
            }
        })
        pinned_used.append({"id": pinned["id"], "ts": ts, "filename": filename})
        img_idx += 1; cue_n += 1
        cur += IMAGE_INTERVAL
        continue

    # ── Normal cue ─────────────────────────────────────────────────────────
    is_video = (cue_n > 0 and cue_n % VIDEO_EVERY == 0)

    if cue_n < len(PROMPT_LIST):
        description, mood = PROMPT_LIST[cue_n]
    else:
        description = f"Wide cinematic shot — ancient Ohio burial mound at golden hour, {n_t[:60]}"
        mood = "ancient, atmospheric"

    # ── Drift capture (for drift_check.xlsx) ───────────────────────────────
    drift_entries.append({
        "cue_n": cue_n, "ts": ts, "sec": cur,
        "type": "VIDEO" if is_video else "IMAGE",
        "narration": n_t,
        "description": description,
    })

    if is_video:
        motion   = MOTION_ROTATION[vid_idx % len(MOTION_ROTATION)]
        filename = f"video_{vid_idx+1:04d}.mp4"
        veo      = f"{motion} — {description} — {STYLE}, {', '.join(PALETTE)}, {LIGHTING} — Mood: {mood}"
        video_prompts.append({
            "index": vid_idx+1, "timestamp": ts,
            "duration_seconds": 6, "filename": filename,
            "script_line": n_t[:80],
            "veo_prompt": veo,
            "nb2_prompt": {
                "shot_type": motion, "subject": description,
                "action": f"camera {motion}, slow deliberate movement",
                "environment": "see subject description",
                "style": STYLE, "color_palette": PALETTE, "lighting": LIGHTING,
                "mood": mood, "negative_constraints": NEG_VID,
            }
        })
        vid_idx += 1
    else:
        shot     = SHOT_ROTATION[img_idx % len(SHOT_ROTATION)]
        filename = f"image_{img_idx+1:04d}.png"
        image_prompts.append({
            "index": img_idx+1, "timestamp": ts,
            "duration": IMAGE_INTERVAL, "filename": filename,
            "script_line": n_t[:80],
            "nb2_prompt": {
                "shot_type": shot, "subject": description,
                "action": "static, slight vignette, subtle depth of field",
                "environment": "see subject description",
                "style": STYLE, "color_palette": PALETTE, "lighting": LIGHTING,
                "mood": mood, "negative_constraints": NEG,
            }
        })
        img_idx += 1

    cue_n += 1
    cur   += IMAGE_INTERVAL

# Re-index
for i, p in enumerate(image_prompts): p["index"]=i+1; p["filename"]=f"image_{i+1:04d}.png"
for i, p in enumerate(video_prompts): p["index"]=i+1; p["filename"]=f"video_{i+1:04d}.mp4"

print(f"✓ Prompts generated: {len(image_prompts)} images, {len(video_prompts)} videos")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Music & SFX cues
# ─────────────────────────────────────────────────────────────────────────────
music_prompts = []
for i, (start, mtype, dur, prompt) in enumerate([
    (0,    "underscore", 55,  "Low ambient documentary underscore — sparse piano, deep cello drone — institutional mystery opening"),
    (45,   "underscore", 105, "Growing investigative tension strings — evidence accumulation builds — slow brooding crescendo"),
    (360,  "peak",       25,  "Orchestral sting — revelation moment"),
    (540,  "underscore", 180, "Somber investigative underscore — archival evidence section — measured gravity"),
    (900,  "underscore", 120, "Reflective documentary scoring — ambiguity and institutional pressure"),
    (1200, "peak",       40,  "Climactic orchestral build — suppression pattern confirmed"),
    (1440, "outro",      132, "Fading unresolved atmospheric outro — open question — tension that will not close"),
], 1):
    music_prompts.append({
        "index": i, "timestamp": sec_to_ts(start),
        "duration_seconds": dur, "type": mtype,
        "prompt": prompt, "filename": f"music_{i:02d}.mp3"
    })

sfx_prompts = []
for i, (start, desc, dur) in enumerate([
    (22,   "deep subterraneal geological rumble — excavation resonance", 4),
    (168,  "old parchment paper rustling and unfolding", 3),
    (340,  "heavy cast iron institutional door closing — metallic reverb", 3),
    (485,  "muffled archival drawer sliding open and shut", 2),
    (740,  "low wind moving through tall grass at burial mound site — desolate", 5),
    (960,  "single typewriter key strike — institutional document moment", 2),
    (1200, "low single bass impact — revelatory moment", 2),
    (1380, "distant crowd ambience fading to silence — historical echo", 4),
    (1492, "single low tone sustained — end of evidence chain — unresolved", 3),
], 1):
    sfx_prompts.append({
        "index": i, "timestamp": sec_to_ts(start),
        "duration_seconds": dur, "description": desc,
        "prompt": desc, "filename": f"sfx_{i:02d}.mp3"
    })

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Write asset_prompts.json
# ─────────────────────────────────────────────────────────────────────────────
result = {
    "TOPIC": TOPIC, "projectDir": PROJ,
    "totalDuration": ACTUAL_DUR, "durationStr": sec_to_ts(ACTUAL_DUR),
    "imageTotalCount": len(image_prompts), "videoTotalCount": len(video_prompts),
    "musicTotalCount": len(music_prompts), "sfxTotalCount": len(sfx_prompts),
    "pinnedClipsUsed": pinned_used,
    "imagePrompts": image_prompts, "videoPrompts": video_prompts,
    "musicPrompts": music_prompts, "sfxPrompts": sfx_prompts,
}
with open(f"{SCRIPTS}/asset_prompts.json", "w") as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
ratio_img_vid = len(image_prompts)/max(1,len(video_prompts))
print(f"✓ asset_prompts.json  {len(image_prompts)} img / {len(video_prompts)} vid = {ratio_img_vid:.2f}:1")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — media_placement.json
# ─────────────────────────────────────────────────────────────────────────────
placement = []
for p in music_prompts:
    placement.append({"type":"music","timestamp":p["timestamp"],"duration_seconds":p["duration_seconds"],
                       "description":p["prompt"],"mood":p["type"],"script_line":""})
for p in sfx_prompts:
    placement.append({"type":"sfx","timestamp":p["timestamp"],"duration_seconds":p["duration_seconds"],
                       "description":p["prompt"],"mood":"ambient","script_line":""})
for p in image_prompts:
    nb2 = p.get("nb2_prompt",{})
    entry = {"type":"image","timestamp":p["timestamp"],"duration_seconds":p["duration"],
             "description":f"{nb2.get('shot_type','')} — {nb2.get('subject','')}",
             "mood":nb2.get("mood",""),"script_line":p.get("script_line",""),
             "filename":p["filename"],"index":p["index"],
             "nb2_prompt":nb2}
    if p.get("pinned"):
        entry["pinned"]=True; entry["pinned_source"]=p["pinned_source"]
        entry["copy_instruction"]=p["copy_instruction"]
    placement.append(entry)
for p in video_prompts:
    entry = {"type":"video","timestamp":p["timestamp"],"duration_seconds":p["duration_seconds"],
             "description":p.get("veo_prompt",""),"mood":p["nb2_prompt"].get("mood",""),
             "script_line":p.get("script_line",""),"filename":p["filename"],
             "index":p["index"],
             "nb2_prompt":p["nb2_prompt"],"veo_prompt":p.get("veo_prompt","")}
    if p.get("pinned"):
        entry["pinned"]=True; entry["pinned_source"]=p["pinned_source"]
        entry["copy_instruction"]=p["copy_instruction"]
    placement.append(entry)
placement.sort(key=lambda x: ts_key(x["timestamp"]))
with open(f"{SCRIPTS}/media_placement.json","w") as f:
    json.dump(placement, f, indent=2, ensure_ascii=False)
print(f"✓ media_placement.json  {len(placement)} entries")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — XLSX files
# ─────────────────────────────────────────────────────────────────────────────
def write_xlsx(path, sheet_name, headers, rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet_name
    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ef = PatternFill("solid", fgColor="D9E2F3")
    for rn, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=rn, column=col, value=val)
            if rn % 2 == 0: c.fill = ef
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers)+1):
        ml = max((len(str(ws.cell(r,col).value or "")) for r in range(1,min(ws.max_row+1,50))), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(ml+2, 80)
    wb.save(path)
    print(f"  ✓ {path.split('/')[-1]}  ({ws.max_row-1} rows)")

def img_prompt_val(p):
    if p.get("pinned"):
        return f'"[PINNED: {p["pinned_source"]}]"'
    nb2 = p.get("nb2_prompt", {})
    return nb2.get("subject", "")

def vid_prompt_val(p):
    if p.get("pinned"):
        return f'"[PINNED CHANNEL ASSET — copy {p["pinned_source"]} → videos/{p["filename"]}]"'
    return p.get("veo_prompt","")

write_xlsx(f"{SCRIPTS}/image_prompts.xlsx","Image Prompts",
    ["Index","Timestamp","ScriptLine","Mood","Prompt","Filename"],
    [[p["index"],p["timestamp"],p.get("script_line",""),
      p.get("nb2_prompt",{}).get("mood",""),
      img_prompt_val(p),f"image_{p['index']:04d}"] for p in image_prompts])

write_xlsx(f"{SCRIPTS}/video_prompts.xlsx","Video Prompts",
    ["Index","Timestamp","ScriptLine","Mood","Prompt","Filename"],
    [[p["index"],p["timestamp"],p.get("script_line",""),
      p["nb2_prompt"].get("mood",""),
      vid_prompt_val(p),f"video_{p['index']:04d}"] for p in video_prompts])

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7b — Drift Check
# For every non-pinned cue, compares the PROMPT_LIST visual description against
# the actual narration at that exact timestamp.  Overlap < 5% → flagged yellow.
# ─────────────────────────────────────────────────────────────────────────────
# ── Drift helpers ─────────────────────────────────────────────────────────────
# Two-signal flagging strategy:
#   1. ACRONYM ANCHOR: if the narrator says a specific acronym (FOIA, NAGPRA,
#      NYT, etc.) the description should reference it. Missing → flag.
#   2. ZERO OVERLAP: both narration and description are content-rich (≥8 tokens
#      each) but share zero words → likely showing wrong section entirely.
# The side-by-side table is the primary tool; flags are attention guides only.

_DRIFT_NOISE = {
    "wide","shot","extreme","medium","long","slow","fast","camera","push","pull",
    "dolly","tilt","static","cinematic","photorealistic","atmospheric","gaslit",
    "amber","ochre","dark","warm","deep","golden","shadow","sidelighting",
    "chiaroscuro","teal","orange","desaturation","vignette","blur","motion",
    "lens","flare","overexposure","flat","elements","limbs","palette","style",
    "subject","action","environment","constraints","subtle","deliberate",
    "foreground","background","resting","captured","framed","exposed","visible",
    "beside","behind","above","below","across","through","toward","corner",
    "left","right","front","back","edge","angle",
}

# Common all-caps words in natural speech that are NOT meaningful acronyms
_COMMON_CAPS = {"THE","AND","BUT","FOR","NOT","ARE","WAS","HAS","HAVE","BEEN",
                "FROM","WITH","THIS","THAT","THEY","WILL","WHAT","WHEN","WHERE",
                "THEN","ALSO","JUST","EVEN","ONLY","OVER","BEEN","INTO","MORE",
                "VERY","THAN","DOES","SAID","WELL","WERE","SOME","LIKE","ONCE",
                # Date/era suffixes — not meaningful acronyms for visual matching
                "BCE","BCE","CE","AD","BC",
                # Common abbreviations whose visual equivalent doesn't need the letters
                "USA","AKA","VIA","EST","ETC",}

def _acronym_anchors(text):
    """Extract ALL-CAPS tokens ≥3 chars that look like true acronyms (not shouting)."""
    found = re.findall(r'\b([A-Z]{3,})\b', text)
    return {w.lower() for w in found if w not in _COMMON_CAPS}

def _content_tokens(text):
    words = re.findall(r"[a-zA-Z0-9]+", text)
    return {w.lower() for w in words if len(w) >= 4 and w.lower() not in _DRIFT_NOISE}

def _overlap(desc, narr):
    d, n = _content_tokens(desc), _content_tokens(narr)
    if not d or not n: return 0.0
    return len(d & n) / len(d | n)

def _drift_flag(desc, narr):
    """Returns (flag_label, reason). Empty strings = no flag."""
    if len(narr.split()) < 6:
        return "", ""   # too short to judge

    # Signal 1: acronym in narration missing from description
    acros = _acronym_anchors(narr)
    if acros:
        desc_l  = desc.lower()
        missing = [a for a in acros if a not in desc_l]
        if missing:
            return "⚠️ DRIFT", f"acronym not in visual: {', '.join(sorted(missing))}"

    # Signal 2: both sides content-rich but zero word overlap
    d_tok = _content_tokens(desc)
    n_tok = _content_tokens(narr)
    if len(d_tok) >= 10 and len(n_tok) >= 10 and len(d_tok & n_tok) == 0:
        return "⚠️ DRIFT", "zero overlap (content-rich)"

    return "", ""

def write_drift_xlsx(path, entries):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Drift Check"
    headers = ["Cue", "Timestamp", "Type",
               "Narration — what speaker says",
               "Visual — what shows on screen",
               "Overlap", "Flag / Missing Entity"]
    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    flag_fill = PatternFill("solid", fgColor="FFD966")   # yellow
    alt_fill  = PatternFill("solid", fgColor="D9E2F3")   # light blue
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    flagged = 0
    for rn, e in enumerate(entries, 2):
        score         = round(_overlap(e["description"], e["narration"]), 3)
        flag, reason  = _drift_flag(e["description"], e["narration"])
        flag_cell     = f"{flag} {reason}".strip() if flag else ""
        if flag: flagged += 1
        row = [e["cue_n"], e["ts"], e["type"],
               e["narration"], e["description"],
               score, flag_cell]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=rn, column=col, value=val)
            if flag:
                c.fill = flag_fill
                c.font = Font(bold=True, size=9)
            elif rn % 2 == 0:
                c.fill = alt_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[rn].height = 48
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 58
    ws.column_dimensions["E"].width = 58
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 30
    ws.freeze_panes = "D2"
    wb.save(path)
    print(f"  ✓ drift_check.xlsx  ({len(entries)} cues, {flagged} flagged ⚠️)")

write_drift_xlsx(f"{SCRIPTS}/drift_check.xlsx", drift_entries)

write_xlsx(f"{SCRIPTS}/music_prompts.xlsx","Music Prompts",
    ["Index","Type","Timestamp","Duration","Prompt","Filename"],
    [[p["index"],p["type"],p["timestamp"],p["duration_seconds"],p["prompt"],p["filename"]] for p in music_prompts])

write_xlsx(f"{SCRIPTS}/sfx_prompts.xlsx","SFX Prompts",
    ["Index","Timestamp","Duration","Prompt","Filename"],
    [[p["index"],p["timestamp"],p["duration_seconds"],p["prompt"],p["filename"]] for p in sfx_prompts])

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7c — Thumbnail Prompts
# Generates 3 thumbnail variants per episode:
#   evidence   — close-up of the key document/artifact (from hook section)
#   site       — dramatic wide shot of the primary archaeological site (mid-episode)
#   atmospheric— symbolic/ominous composition (from climax section)
#
# Each variant includes:
#   - NB2 image prompt (hero visual, text-safe top zone, style-locked)
#   - 3 text overlay suggestions (ALL CAPS, ≤7 words, suppression register)
#   - Composition notes for the designer
# ─────────────────────────────────────────────────────────────────────────────

def _thumb_text_suggestions(topic):
    """
    Generate 3 ALL-CAPS thumbnail text lines from the episode topic.
    Targets: ≤7 words, conspiracy/suppression register.

    Strategy: extract the strongest VERB and the strongest NOUN from the topic,
    then build three patterns around them.
    Returns list of 3 strings.
    """
    title_part = topic.split(":")[0].strip().upper()
    words = [w.strip(",.;:?!'\"") for w in title_part.split()]

    # Known suppression/action verbs — pick the first one found in the title
    STRONG_VERBS = ["DESTROYED","BURIED","ERASED","HID","REMOVED","SUPPRESSED",
                    "CONCEALED","COVERED","DENIED","BANNED","LOST","VANISHED"]
    FILLER       = {"THE","A","AN","OF","IN","ON","AT","TO","IS","ARE","WAS","WERE",
                    "AND","OR","BUT","BY","FOR","AS","WITH","FROM","THEIR","THEY",
                    "THIS","THAT","IT","ITS","BE","HAS","HAVE","HAD","NOT","THEIR",
                    "OWN","PROVE","RECORDS"}

    found_verb  = next((w for w in words if w in STRONG_VERBS), None)
    verb_idx    = words.index(found_verb) if found_verb else -1
    all_nouns   = [w for w in words if w not in FILLER and w not in STRONG_VERBS and len(w) >= 4]

    # Nouns AFTER the verb are the object (what was destroyed/hidden)
    # Nouns BEFORE the verb are the actor (who did it)
    obj_nouns   = [w for w in words[verb_idx+1:] if w not in FILLER and w not in STRONG_VERBS and len(w) >= 4] if verb_idx >= 0 else all_nouns
    act_nouns   = [w for w in words[:verb_idx]  if w not in FILLER and w not in STRONG_VERBS and len(w) >= 4] if verb_idx >= 0 else []

    verb    = found_verb or "HID"
    obj     = obj_nouns[0]  if obj_nouns  else (all_nouns[0] if all_nouns else "EVIDENCE")
    actor   = act_nouns[-1] if act_nouns  else "THEY"   # last word before verb = institution name

    suggestions = [
        f"THEY {verb} THE {obj}",            # "THEY DESTROYED THE GIANTS"
        f"THE {obj} THEY {verb}",            # "THE GIANTS THEY DESTROYED"
        f"{actor} {verb} THE {obj}",         # "SMITHSONIAN DESTROYED THE GIANTS"
    ]
    # Cap each at 7 words
    suggestions = [" ".join(s.split()[:7]) for s in suggestions]
    return suggestions

def _thumb_nb2_prompt(variant, description, topic, tcfg):
    """
    Build a full 12-field NB2 JSON prompt for a thumbnail variant.
    Schema: goal, subject, context, style, composition, lighting,
            color_palette, background, camera_or_lens, mood,
            text_space, negative_constraints
    """
    hero  = tcfg.get("hero_visual", {})
    tover = tcfg.get("text_overlay", {})
    neg   = hero.get("negative", [])
    zone_pct = tover.get("zone_height_pct", 22)

    # Variant-specific camera and composition settings
    variant_cfg = {
        "evidence": {
            "composition": f"extreme close-up — subject fills lower 78% of frame; top {zone_pct}% intentionally underexposed for text overlay",
            "camera_or_lens": {"focal_length": "85mm", "aperture": "f/2.8", "type": "DSLR"},
            "context": "dark archive desk or institutional shelf — shallow depth of field",
            "background": "deep shadow receding into darkness, faint archival texture, bokeh",
            "goal": f"Thumbnail hero — close-up of primary documentary evidence for: {topic}",
        },
        "site": {
            "composition": f"wide establishing shot — strong foreground subject, layered depth; top {zone_pct}% intentionally darker sky for text overlay",
            "camera_or_lens": {"focal_length": "24mm", "aperture": "f/5.6", "type": "DSLR"},
            "context": "active archaeological excavation or burial mound landscape",
            "background": "dramatic layered horizon, atmospheric depth, amber sky",
            "goal": f"Thumbnail hero — wide shot of primary archaeological site for: {topic}",
        },
        "atmospheric": {
            "composition": f"wide silhouette shot — subject dark against sky; top {zone_pct}% open sky intentionally darker for text overlay",
            "camera_or_lens": {"focal_length": "16mm", "aperture": "f/8", "type": "DSLR"},
            "context": "ancient burial mound or earthwork at twilight or dusk",
            "background": "dramatic twilight sky, deep amber and indigo tones, atmospheric haze",
            "goal": f"Thumbnail hero — ominous atmospheric wide shot for: {topic}",
        },
    }
    vcfg = variant_cfg.get(variant, variant_cfg["site"])

    return {
        "goal":         vcfg["goal"],
        "subject":      [description],
        "context":      vcfg["context"],
        "style":        hero.get("style", "Photorealistic, cinematic, 8K, ultra-sharp, extreme clarity"),
        "composition":  vcfg["composition"],
        "lighting":     hero.get("lighting", LIGHTING),
        "color_palette": hero.get("color_palette", PALETTE),
        "background":   vcfg["background"],
        "camera_or_lens": vcfg["camera_or_lens"],
        "mood":         "ancient, mysterious, slightly ominous — thumbnail impact level",
        "text_space":   f"top-third — top {zone_pct}% of frame reserved for distressed ALL-CAPS text overlay; do not place key subject detail there",
        "negative_constraints": neg,
    }

def _pick_thumb_cues():
    """
    Pick 3 source cues from image_prompts for thumbnail hero visuals:
      evidence   — first document/evidence-heavy cue in the hook section (cues 0-15%)
      site       — most visually striking wide-shot cue near mid-episode (45-65%)
      atmospheric— an ominous atmospheric cue near the climax (75-90%)
    Returns dict with variant → (description, timestamp)
    """
    total = len(image_prompts)
    if total == 0:
        return {}

    def pick_in_range(lo_pct, hi_pct, keywords=None):
        lo = int(total * lo_pct); hi = int(total * hi_pct)
        candidates = image_prompts[lo:hi]
        if keywords:
            kw_lower = [k.lower() for k in keywords]
            ranked = sorted(candidates,
                key=lambda p: sum(k in p.get("nb2_prompt",{}).get("subject","").lower()
                                  for k in kw_lower), reverse=True)
            return ranked[0] if ranked else (candidates[0] if candidates else None)
        return candidates[len(candidates)//2] if candidates else None

    evidence_cue    = pick_in_range(0.0, 0.20,
        ["report","memo","document","record","annual","1868","1894","smithsonian"])
    site_cue        = pick_in_range(0.35, 0.60,
        ["mound","burial","excavation","earthwork","site","ohio","virginia","illinois"])
    atmospheric_cue = pick_in_range(0.88, 1.00,
        ["shadow","silhouette","horizon","dusk","twilight","vast","sky","landscape",
         "mound","earthwork","golden","ohio","burial","ancient"])

    result = {}
    if evidence_cue:
        desc = evidence_cue.get("nb2_prompt", {}).get("subject", "")
        result["evidence"] = (desc, evidence_cue.get("timestamp",""))
    if site_cue:
        desc = site_cue.get("nb2_prompt", {}).get("subject", "")
        result["site"] = (desc, site_cue.get("timestamp",""))
    if atmospheric_cue:
        desc = atmospheric_cue.get("nb2_prompt", {}).get("subject", "")
        result["atmospheric"] = (desc, atmospheric_cue.get("timestamp",""))
    return result

def write_thumbnail_xlsx(path, topic, tcfg):
    if not tcfg:
        return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Thumbnail Prompts"

    headers = [
        "Variant", "Source Timestamp",
        "Text Overlay — Line 1", "Text Overlay — Line 2", "Text Overlay — Line 3",
        "NB2 Full JSON Prompt",
        "goal", "subject", "context", "style", "composition",
        "lighting", "color_palette", "background",
        "camera_focal_length", "camera_aperture", "camera_type",
        "mood", "text_space", "negative_constraints",
        "Branding Note",
    ]
    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    var_fills = {
        "evidence":    PatternFill("solid", fgColor="C6EFCE"),  # green
        "site":        PatternFill("solid", fgColor="DDEBF7"),  # blue
        "atmospheric": PatternFill("solid", fgColor="FCE4D6"),  # orange
    }
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    thumb_cues = _pick_thumb_cues()
    text_lines = _thumb_text_suggestions(topic)
    branding   = tcfg.get("branding", {})
    brand_note = f"'{branding.get('channel_name','RUINS UNTOLD')}' — {branding.get('style','distressed stamp')}, {branding.get('position','bottom-right')}, {branding.get('color','#F5E6C8')}, {int(branding.get('opacity', 0.6)*100)}% opacity — add in post."

    for rn, variant in enumerate(["evidence", "site", "atmospheric"], 2):
        if variant not in thumb_cues:
            continue
        desc, src_ts = thumb_cues[variant]
        nb2 = _thumb_nb2_prompt(variant, desc, topic, tcfg)
        cam = nb2["camera_or_lens"]

        row = [
            variant.upper(),
            src_ts,
            text_lines[0] if len(text_lines) > 0 else "",
            text_lines[1] if len(text_lines) > 1 else "",
            text_lines[2] if len(text_lines) > 2 else "",
            json.dumps(nb2, indent=2),           # full NB2 JSON — paste directly into generator
            nb2["goal"],
            "\n".join(nb2["subject"]),
            nb2["context"],
            nb2["style"],
            nb2["composition"],
            nb2["lighting"],
            ", ".join(nb2["color_palette"]),
            nb2["background"],
            cam.get("focal_length", ""),
            cam.get("aperture", ""),
            cam.get("type", ""),
            nb2["mood"],
            nb2["text_space"],
            ", ".join(nb2["negative_constraints"]),
            brand_note,
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill = var_fills.get(variant, PatternFill())
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[rn].height = 100

    # Column widths
    widths = [14, 12, 28, 28, 28,  # variant, ts, text lines
              55,                   # full JSON
              50, 55, 35, 28, 55,  # goal, subject, context, style, composition
              32, 32, 40,           # lighting, palette, background
              14, 12, 12,           # camera fields
              32, 50, 55,           # mood, text_space, negative
              45]                   # branding
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(path)
    print(f"  ✓ thumbnail_prompts.xlsx  (3 variants — evidence / site / atmospheric)")

write_thumbnail_xlsx(f"{SCRIPTS}/thumbnail_prompts.xlsx", TOPIC, THUMB_CFG)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 — assemble.sh with FIXED timing (scaled durations, zero drift)
# ─────────────────────────────────────────────────────────────────────────────
visuals = []
for p in image_prompts:
    visuals.append({"type":"image","index":p["index"],"ts":ts_key(p["timestamp"]),
                    "filename":p["filename"],"base_dur":float(p["duration"]),
                    "pinned_source":p.get("pinned_source")})
for p in video_prompts:
    visuals.append({"type":"video","index":p["index"],"ts":ts_key(p["timestamp"]),
                    "filename":p["filename"],"base_dur":float(p["duration_seconds"]),
                    "pinned_source":p.get("pinned_source")})
visuals.sort(key=lambda x:(x["ts"], 0 if x["type"]=="video" else 1))

total_base   = sum(v["base_dur"] for v in visuals)
ratio        = ACTUAL_DUR / total_base
for v in visuals:
    v["scaled_dur"] = v["base_dur"] * ratio
total_scaled = sum(v["scaled_dur"] for v in visuals)

IMG_DIR_  = f"{PROJ}/images"
VID_DIR_  = f"{PROJ}/videos"
AUDIO_DIR = f"{PROJ}/audio"
MUSIC_DIR = f"{PROJ}/music"
CLIPS_DIR = f"{PROJ}/clips"
OUTPUT    = f"{PROJ.rstrip('/')}_final.mp4"

sh = [
    "#!/bin/sh","set -e","",
    f'PROJECT_DIR="{PROJ}"', f'OUTPUT="{OUTPUT}"', f'CLIPS_DIR="{CLIPS_DIR}"',
    f'AUDIO="{AUDIO_DIR}/voiceover.mp3"',"",
    'mkdir -p "$CLIPS_DIR"',"",
    f"# {len(visuals)} clips | ratio={ratio:.6f} | drift={total_scaled-ACTUAL_DUR:+.4f}s",
    f"# image_dur={7.0*ratio:.4f}s  video_dur={6.0*ratio:.4f}s",
    "","# ── Clips ───────────────────────────────────────────────────────────────────",
]

for v in visuals:
    sdur  = v["scaled_dur"]
    fo_st = sdur - FADE_DUR
    clip  = f'{CLIPS_DIR}/clip_{v["ts"]:05d}_{v["index"]:04d}.mp4'
    scale_pad = ("scale=1920:1080:force_original_aspect_ratio=decrease,"
                 "pad=1920:1080:(ow-iw)/2:(oh-ih)/2,setsar=1")
    fades = f"fade=t=in:st=0:d={FADE_DUR},fade=t=out:st={fo_st:.4f}:d={FADE_DUR}"

    if v.get("pinned_source"):
        src = v["pinned_source"]
        sh.append(f'# PINNED: {src}')
        sh.append(f'ffmpeg -y -i "{src}" -t {sdur:.4f} -vf "{scale_pad},{fades}" -r 24 -c:v libx264 -pix_fmt yuv420p "{clip}"')
    elif v["type"] == "image":
        src = f'{IMG_DIR_}/{v["filename"].replace(".png", ".jpg")}'
        sh.append(f'ffmpeg -y -loop 1 -i "{src}" -t {sdur:.4f} -vf "{scale_pad},{fades}" -r 24 -c:v libx264 -pix_fmt yuv420p "{clip}"')
    else:
        src = f'{VID_DIR_}/{v["filename"]}'
        sh.append(f'ffmpeg -y -i "{src}" -t {sdur:.4f} -vf "{scale_pad},setpts=PTS*{ratio:.6f},{fades}" -r 24 -c:v libx264 -pix_fmt yuv420p "{clip}"')

sh += [
    "","# ── Concat ──────────────────────────────────────────────────────────────────",
    'CONCAT_LIST="$PROJECT_DIR/scripts/concat.txt"','> "$CONCAT_LIST"',
]
for v in visuals:
    sh.append(f'echo "file \'$CLIPS_DIR/clip_{v["ts"]:05d}_{v["index"]:04d}.mp4\'" >> "$CONCAT_LIST"')

sh += [
    "","# ── Music ───────────────────────────────────────────────────────────────────",
    f'MUSIC_LIST="{PROJ}/scripts/music_concat.txt"','> "$MUSIC_LIST"',
]
for p in music_prompts:
    sh.append(f'echo "file \'{MUSIC_DIR}/{p["filename"]}\'" >> "$MUSIC_LIST"')
sh.append(f'ffmpeg -y -f concat -safe 0 -i "$MUSIC_LIST" -c copy "{AUDIO_DIR}/music_combined_raw.mp3"')
sh.append(f'ffmpeg -y -i "{AUDIO_DIR}/music_combined_raw.mp3" -af loudnorm=I=-23:TP=-2:LRA=7 "{AUDIO_DIR}/music_combined.mp3"')

sh += [
    "","# ── Final assembly ──────────────────────────────────────────────────────────",
    f'SILENT_VIDEO="{PROJ}/scripts/silent_video.mp4"',
    f'ffmpeg -y -f concat -safe 0 -i "$CONCAT_LIST" -c copy "$SILENT_VIDEO"',
    "","# Normalize voiceover to -16 LUFS broadcast standard",
    f'VOICEOVER_NORM="{AUDIO_DIR}/voiceover_norm.mp3"',
    f'ffmpeg -y -i "$AUDIO" -af loudnorm=I=-16:TP=-1.5:LRA=11 "$VOICEOVER_NORM"',
    "","# Mix normalized voiceover (100%) + music (25%)",
    f'MIXED_AUDIO="{AUDIO_DIR}/mixed_audio.mp3"',
    f'ffmpeg -y -i "$VOICEOVER_NORM" -i "{AUDIO_DIR}/music_combined.mp3" \\',
    f'  -filter_complex "[0:a]volume=\'if(lt(t,618),1.0,0.50)\':eval=frame[vo];[1:a]volume=0.25[mu];[vo][mu]amix=inputs=2:duration=first[out]" \\',
    f'  -map "[out]" "$MIXED_AUDIO"',
    "","# Mux video + mixed audio",
    f'ffmpeg -y -i "$SILENT_VIDEO" -i "$MIXED_AUDIO" -c:v copy -c:a aac -shortest "$OUTPUT"',
    "",f'echo "✓ Done: $OUTPUT"',
]

asm_path = f"{SCRIPTS}/assemble.sh"
with open(asm_path,"w") as f:
    f.write("\n".join(sh)+"\n")
os.chmod(asm_path, 0o755)
print(f"✓ assemble.sh  ({len(visuals)} clips, ratio={ratio:.6f}, drift={total_scaled-ACTUAL_DUR:+.4f}s)")

# ─────────────────────────────────────────────────────────────────────────────
# Summary
# ─────────────────────────────────────────────────────────────────────────────
print()
print("═"*60)
print(f"  Topic       : {TOPIC}")
print(f"  Images      : {len(image_prompts)}")
print(f"  Videos      : {len(video_prompts)}")
print(f"  Duration    : {sec_to_ts(ACTUAL_DUR)}")
print(f"  Ratio       : {ratio:.6f}")
print(f"  Drift       : {total_scaled-ACTUAL_DUR:+.6f}s")
print(f"  Prompt mode : manually written from transcript (v4 — no extraction)")
print("═"*60)

if pinned_used:
    print()
    print("ACTION REQUIRED — copy pinned assets before running assemble.sh:")
    for pu in pinned_used:
        for vp in video_prompts:
            if vp.get("pinned_id") == pu["id"] and vp["filename"] == pu["filename"]:
                print(f"  {vp['copy_instruction']}")
